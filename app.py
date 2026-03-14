"""
FurnitureFlow – Expert-Defined Sequential Loading & Integrity System
Tüm özellikler: Login/Rol, SLA, Öncelik, Barkod, Stok Durumu,
Hasar Yönetimi, Personel Performans, Talep Tahmini, Müşteri Takip
Çalıştır: streamlit run app.py
"""
import sqlite3, io, os, hashlib, math, secrets, smtplib, time, subprocess
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ── Streamlit Cloud: veritabanı yoksa otomatik kurulum ───────────────────────
if not os.path.exists("furnitureflow.db"):
    subprocess.run(["python", "init_db.py"],        check=False)
    subprocess.run(["python", "generate_sample_data.py", "25"], check=False)

import pandas as pd
import streamlit as st
import qrcode
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="FurnitureFlow", page_icon="🪑",
                   layout="wide", initial_sidebar_state="expanded")
DB = "furnitureflow.db"

# ── Font ──────────────────────────────────────────────────────────────────────
def _reg_fonts():
    d = r"C:\Windows\Fonts"
    try:
        pdfmetrics.registerFont(TTFont("Arial",      os.path.join(d,"arial.ttf")))
        pdfmetrics.registerFont(TTFont("Arial-Bold", os.path.join(d,"arialbd.ttf")))
        return "Arial","Arial-Bold"
    except:
        return "Helvetica","Helvetica-Bold"
FN, FB = _reg_fonts()

# ── DB helpers ─────────────────────────────────────────────────────────────────
def get_conn():
    c = sqlite3.connect(DB, check_same_thread=False)
    c.row_factory = sqlite3.Row
    return c
def q(sql,p=()):
    with get_conn() as c: return [dict(r) for r in c.execute(sql,p).fetchall()]
def q1(sql,p=()):
    with get_conn() as c:
        r = c.execute(sql,p).fetchone(); return dict(r) if r else None
def ex(sql,p=()):
    with get_conn() as c: c.execute(sql,p); c.commit()
def ex_id(sql,p=()):
    with get_conn() as c:
        cur=c.execute(sql,p); c.commit(); return cur.lastrowid

# ── Auth ───────────────────────────────────────────────────────────────────────
def sha256(pw): return hashlib.sha256(pw.encode()).hexdigest()
def login(u,p): return q1("SELECT * FROM users WHERE username=? AND password_hash=? AND is_active=1",(u,sha256(p)))
def can(action):
    role = st.session_state.get("user",{}).get("role","")
    PERMS = {
        "new_order":    ["admin","yonetici"],
        "completed":    ["admin","yonetici"],
        "analytics":    ["admin","yonetici"],
        "stock":        ["admin","yonetici","personel"],
        "admin":        ["admin"],
        "notifications":["admin","yonetici"],
    }
    return role in PERMS.get(action, ["admin","yonetici","personel"])

# ── SLA helpers ────────────────────────────────────────────────────────────────
def sla_status(order):
    """Returns (elapsed_min, sla_min, pct, label, color)"""
    if not order.get("started_at"): return 0, order["sla_minutes"], 0, "Başlamadı", "#6b7280"
    start = datetime.strptime(order["started_at"], "%Y-%m-%d %H:%M:%S")
    if order.get("completed_at"):
        end = datetime.strptime(order["completed_at"], "%Y-%m-%d %H:%M:%S")
    else:
        end = datetime.now()
    elapsed = (end - start).total_seconds() / 60
    sla     = order["sla_minutes"] or 90
    pct     = min(elapsed / sla, 1.0)
    if pct < 0.6:    label, color = "✅ Normal",  "#22c55e"
    elif pct < 0.85: label, color = "⚠️ Dikkat",  "#f97316"
    else:            label, color = "🚨 Kritik",  "#ef4444"
    return round(elapsed,1), sla, pct, label, color

def priority_score(order):
    scores = {"urgent":3,"high":2,"normal":1,"low":0}
    base = scores.get(order.get("priority","normal"),1)
    # add penalty for SLA breach
    elapsed, sla, pct, _, _ = sla_status(order)
    return base + (2 if pct >= 1.0 else 1 if pct >= 0.85 else 0)

# ── QR ────────────────────────────────────────────────────────────────────────
def make_qr(data:str) -> bytes:
    qr = qrcode.QRCode(version=1,box_size=6,border=3)
    qr.add_data(data); qr.make(fit=True)
    img = qr.make_image(fill_color="#1a1a26",back_color="white")
    buf = io.BytesIO(); img.save(buf,"PNG"); buf.seek(0); return buf.read()

# ── Customer token ─────────────────────────────────────────────────────────────
def get_or_create_token(order_id:int) -> str:
    row = q1("SELECT token FROM customer_tokens WHERE order_id=?", (order_id,))
    if row: return row["token"]
    token = secrets.token_urlsafe(16)
    ex("INSERT INTO customer_tokens (order_id,token) VALUES(?,?)", (order_id,token))
    return token

# ── Vehicle capacity ──────────────────────────────────────────────────────────
def calc_vehicles(m3:float):
    vs = q("SELECT name,capacity_m3 FROM vehicle_types ORDER BY capacity_m3 DESC")
    out = []
    for v in vs:
        cnt  = math.ceil(m3/v["capacity_m3"]) if m3>0 else 0
        fill = (m3/(cnt*v["capacity_m3"])*100) if cnt>0 else 0
        out.append({"name":v["name"],"capacity":v["capacity_m3"],"count":cnt,"fill":fill})
    return out

# ── Notifications ─────────────────────────────────────────────────────────────
def push_notif(type_:str, msg:str, order_id=None):
    ex("INSERT INTO notifications (type,message,order_id) VALUES(?,?,?)", (type_,msg,order_id))

def audit(action:str, entity:str, entity_id=None, detail:str=""):
    user = st.session_state.get("user",{})
    username = user.get("username","system") if user else "system"
    ex("INSERT INTO audit_log (username,action,entity,entity_id,detail) VALUES(?,?,?,?,?)",
       (username, action, entity, str(entity_id) if entity_id else None, detail))

def unread_count():
    r = q1("SELECT COUNT(*) c FROM notifications WHERE is_read=0")
    return r["c"] if r else 0

# ── Email sender ──────────────────────────────────────────────────────────────
def send_email(to:str, subject:str, body:str, pdf_bytes:bytes=None, pdf_name:str="packing_list.pdf") -> bool:
    cfg = {r["key"]:r["value"] for r in q("SELECT key,value FROM app_settings WHERE key LIKE 'smtp%'")}
    if not cfg.get("smtp_user") or not cfg.get("smtp_pass"): return False
    try:
        msg = MIMEMultipart()
        msg["From"] = cfg["smtp_user"]; msg["To"] = to; msg["Subject"] = subject
        msg.attach(MIMEText(body,"html"))
        if pdf_bytes:
            part = MIMEBase("application","octet-stream")
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition",f'attachment; filename="{pdf_name}"')
            msg.attach(part)
        with smtplib.SMTP(cfg.get("smtp_host","smtp.gmail.com"), int(cfg.get("smtp_port",587))) as s:
            s.starttls(); s.login(cfg["smtp_user"],cfg["smtp_pass"]); s.send_message(msg)
        return True
    except Exception as e:
        st.error(f"E-posta gönderilemedi: {e}"); return False

# ── PDF generator ─────────────────────────────────────────────────────────────
def make_pdf(order:dict, items:list) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf,pagesize=A4,leftMargin=2*cm,rightMargin=2*cm,
                            topMargin=2*cm,bottomMargin=2*cm)
    T  = ParagraphStyle("t",fontSize=18,leading=22,alignment=TA_CENTER,fontName=FB,spaceAfter=4)
    S  = ParagraphStyle("s",fontSize=10,alignment=TA_CENTER,fontName=FN,
                        textColor=colors.HexColor("#555555"),spaceAfter=12)
    SL = {"normal":"✓ Normal","damaged":"⚠ Hasarlı","missing":"✗ Eksik"}
    PL = {"normal":"Normal","urgent":"ACİL","high":"Yüksek","low":"Düşük"}
    el = []
    el.append(Paragraph("PACKING LIST / ÇEKİ LİSTESİ",T))
    el.append(Paragraph("FurnitureFlow · Luxe Life Mobilya",S))
    el.append(HRFlowable(width="100%",thickness=1.5,color=colors.HexColor("#7c6dff")))
    el.append(Spacer(1,.4*cm))
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    elapsed,sla,_,_,_ = sla_status(order)
    info = [
        ["Sipariş No",  order.get("order_number","")],
        ["Koleksiyon",  order.get("collection","")],
        ["Öncelik",     PL.get(order.get("priority","normal"),"Normal")],
        ["Müşteri",     order.get("customer_name") or "—"],
        ["E-posta",     order.get("customer_email") or "—"],
        ["Telefon",     order.get("customer_phone") or "—"],
        ["Hazırlık S.", f"{elapsed:.0f} dk / {sla} dk SLA"],
        ["Tarih",       now],
    ]
    it = Table(info,colWidths=[4*cm,12*cm])
    it.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),FN),("FONTNAME",(0,0),(0,-1),FB),
        ("FONTSIZE",(0,0),(-1,-1),10),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.white,colors.HexColor("#f9f9f9")]),
        ("GRID",(0,0),(-1,-1),.3,colors.HexColor("#e5e7eb")),
        ("PADDING",(0,0),(-1,-1),6),
    ]))
    el.append(it); el.append(Spacer(1,.5*cm))
    header = ["#","Parça Adı","G (cm)","D (cm)","Y (cm)","Hacim (m³)","Durum"]
    td = [header]; tv = 0.0
    for it2 in items:
        v = (it2["width_cm"]*it2["depth_cm"]*it2["height_cm"])/1_000_000; tv+=v
        td.append([str(it2["loading_order"]),it2["name"],
                   str(it2["width_cm"]),str(it2["depth_cm"]),str(it2["height_cm"]),
                   f"{v:.4f}",SL.get(it2.get("item_status") or "normal","✓ Normal")])
    td.append(["","TOPLAM","","","",f"{tv:.4f}",""])
    cw=[.8*cm,5.5*cm,1.9*cm,1.9*cm,1.9*cm,2.3*cm,2.7*cm]
    pt=Table(td,colWidths=cw,repeatRows=1)
    sc=[
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#7c6dff")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),FB),("FONTSIZE",(0,0),(-1,0),8),
        ("ALIGN",(0,0),(-1,0),"CENTER"),
        ("FONTNAME",(0,1),(-1,-2),FN),("FONTSIZE",(0,1),(-1,-2),8),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white,colors.HexColor("#f3f4f6")]),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#f0fdf4")),
        ("FONTNAME",(0,-1),(-1,-1),FB),
        ("GRID",(0,0),(-1,-1),.4,colors.HexColor("#e5e7eb")),
        ("ALIGN",(0,0),(0,-1),"CENTER"),("ALIGN",(2,0),(6,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("PADDING",(0,0),(-1,-1),4),
    ]
    for i,row in enumerate(items,1):
        s=row.get("item_status") or "normal"
        if s=="damaged": sc.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#fff7ed")))
        elif s=="missing": sc.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#fef2f2")))
    pt.setStyle(TableStyle(sc)); el.append(pt); el.append(Spacer(1,.5*cm))
    # vehicle
    vs = calc_vehicles(tv)
    vd = [["Araç Tipi","Kapasite","Adet","Doluluk"]]
    for v in vs:
        vd.append([v["name"],f"{v['capacity']} m³",
                   str(v["count"]) if v["count"]>0 else "—",
                   f"%{v['fill']:.0f}" if v["count"]>0 else "—"])
    vt=Table(vd,colWidths=[4*cm,3*cm,4*cm,5*cm])
    vt.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1e1e2e")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,-1),FN),("FONTNAME",(0,0),(-1,0),FB),
        ("FONTSIZE",(0,0),(-1,-1),9),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#e5e7eb")),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("PADDING",(0,0),(-1,-1),5),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f3f4f6")]),
    ]))
    el.append(Paragraph("Araç Kapasitesi",ParagraphStyle("vc",fontSize=10,fontName=FB,spaceAfter=4,spaceBefore=4)))
    el.append(vt); el.append(Spacer(1,.8*cm))
    sd=[["Hazırlayan","Kontrol Eden","Onaylayan"],["\n\n______________"]*3]
    st2=Table(sd,colWidths=[5.5*cm]*3)
    st2.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,0),FB),("FONTNAME",(0,1),(-1,1),FN),
        ("FONTSIZE",(0,0),(-1,-1),9),("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("GRID",(0,0),(-1,-1),.3,colors.HexColor("#d1d5db")),("PADDING",(0,0),(-1,-1),8),
    ]))
    el.append(st2)
    doc.build(el); buf.seek(0); return buf.read()

# ── Excel generator ───────────────────────────────────────────────────────────
def make_xlsx(order:dict, items:list) -> bytes:
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Çeki Listesi"
    PURPLE="7C6DFF"; DARK="1E1E2E"; LIGHT="F3F4F6"; GREEN="F0FDF4"; ORANGE="FFF7ED"; RED="FEF2F2"
    def hf(h): return PatternFill("solid",fgColor=h)
    def bf(sz=10,wh=False,bold=True): return Font(name="Arial",size=sz,bold=bold,color="FFFFFF" if wh else "000000")
    thin=Side(style="thin",color="E5E7EB"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
    lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
    ws.merge_cells("A1:H1"); ws["A1"]="PACKING LIST / ÇEKİ LİSTESİ"
    ws["A1"].font=Font(name="Arial",size=16,bold=True,color=PURPLE); ws["A1"].alignment=ctr; ws.row_dimensions[1].height=28
    ws.merge_cells("A2:H2"); ws["A2"]="FurnitureFlow · Luxe Life Mobilya"
    ws["A2"].font=Font(name="Arial",size=10,color="555555"); ws["A2"].alignment=ctr
    SL={"normal":"✓ Normal","damaged":"⚠ Hasarlı","missing":"✗ Eksik"}
    PL={"normal":"Normal","urgent":"ACİL","high":"Yüksek","low":"Düşük"}
    elapsed,sla,_,_,_=sla_status(order)
    info=[("Sipariş No",order.get("order_number","")),("Koleksiyon",order.get("collection","")),
          ("Öncelik",PL.get(order.get("priority","normal"),"Normal")),
          ("Müşteri",order.get("customer_name") or "—"),("E-posta",order.get("customer_email") or "—"),
          ("Telefon",order.get("customer_phone") or "—"),
          ("Hazırlık",f"{elapsed:.0f} dk / {sla} dk SLA"),
          ("Tarih",datetime.now().strftime("%d.%m.%Y %H:%M"))]
    for i,(lbl,val) in enumerate(info,4):
        ws.cell(i,1,lbl).font=bf(10); ws.cell(i,1).fill=hf(LIGHT); ws.cell(i,1).alignment=lft; ws.cell(i,1).border=bd
        ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=8)
        ws.cell(i,2,val).font=Font(name="Arial",size=10); ws.cell(i,2).alignment=lft; ws.cell(i,2).border=bd
    heads=["#","Parça Adı","G (cm)","D (cm)","Y (cm)","Hacim (m³)","Durum","Not"]
    cws=[6,34,10,10,10,14,14,20]
    hr=13
    for j,(h,w) in enumerate(zip(heads,cws),1):
        c=ws.cell(hr,j,h); c.font=bf(10,True); c.fill=hf(PURPLE); c.alignment=ctr; c.border=bd
        ws.column_dimensions[get_column_letter(j)].width=w
    ws.row_dimensions[hr].height=20
    tv=0.0
    for idx,it in enumerate(items,hr+1):
        v=(it["width_cm"]*it["depth_cm"]*it["height_cm"])/1_000_000; tv+=v
        s=it.get("item_status") or "normal"
        rf=hf(ORANGE) if s=="damaged" else hf(RED) if s=="missing" else hf("FFFFFF" if idx%2==0 else LIGHT)
        row=[it["loading_order"],it["name"],it["width_cm"],it["depth_cm"],it["height_cm"],round(v,4),SL[s],it.get("item_note") or ""]
        for j,val in enumerate(row,1):
            c=ws.cell(idx,j,val); c.font=Font(name="Arial",size=9); c.fill=rf
            c.alignment=ctr if j!=2 else lft; c.border=bd
    tr=hr+len(items)+1
    for j,val in [(2,"TOPLAM"),(6,round(tv,4))]:
        c=ws.cell(tr,j,val); c.font=bf(10); c.fill=hf(GREEN); c.alignment=ctr; c.border=bd
    for j in [1,3,4,5,7,8]:
        ws.cell(tr,j).fill=hf(GREEN); ws.cell(tr,j).border=bd
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

# ── Route Optimization ────────────────────────────────────────────────────────
def haversine(lat1, lon1, lat2, lon2) -> float:
    """İki koordinat arası mesafe (km)"""
    R = 6371
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

def nearest_neighbor_route(stops: list, depot_lat=40.1826, depot_lng=29.0665) -> tuple:
    """
    Greedy nearest-neighbor TSP.
    stops: [{"id":..,"name":..,"lat":..,"lng":..,"address":..}, ...]
    depot: Bursa merkez (varsayılan)
    Returns: (ordered_stops, total_km)
    """
    if not stops: return [], 0
    unvisited = stops.copy()
    route     = []
    cur_lat, cur_lng = depot_lat, depot_lng
    total_km  = 0.0
    while unvisited:
        nearest = min(unvisited, key=lambda s: haversine(cur_lat, cur_lng, s["lat"], s["lng"]))
        d = haversine(cur_lat, cur_lng, nearest["lat"], nearest["lng"])
        total_km += d
        cur_lat, cur_lng = nearest["lat"], nearest["lng"]
        route.append({**nearest, "dist_from_prev": round(d,1)})
        unvisited.remove(nearest)
    # Return to depot
    total_km += haversine(cur_lat, cur_lng, depot_lat, depot_lng)
    return route, round(total_km, 1)

# Bursa ilçe koordinatları (yaklaşık merkez noktalar)
BURSA_DISTRICTS = {
    "Osmangazi":   (40.1826, 29.0665),
    "Nilüfer":     (40.2131, 28.9769),
    "Yıldırım":    (40.1908, 29.1147),
    "Gemlik":      (40.4319, 29.1536),
    "Mudanya":     (40.3747, 28.8839),
    "Bursa Merkez":(40.1826, 29.0665),
    "İnegöl":      (40.0785, 29.5124),
    "Mustafakemalpaşa":(40.0333, 28.4000),
    "Karacabey":   (40.2167, 28.3500),
    "Gürsu":       (40.2167, 29.1833),
    "Kestel":      (40.1833, 29.2167),
    "Diğer":       (40.1826, 29.0665),
}

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ═══════════════════════════════════════════════════════════════════════════
   TEMEL & RESET
═══════════════════════════════════════════════════════════════════════════ */
*, *::before, *::after { box-sizing: border-box; }

.stApp { background-color: #0d0d12; color: #e2e2ee; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; }

/* Ana içerik genişliği */
.main .block-container {
    max-width: 960px !important;
    padding: 1.5rem 1.5rem 3rem !important;
    margin: 0 auto;
}

/* ═══════════════════════════════════════════════════════════════════════════
   SİDEBAR
═══════════════════════════════════════════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #13131c 0%, #0f0f18 100%) !important;
    border-right: 1px solid #1e1e2e !important;
}
[data-testid="stSidebar"] .stRadio label {
    font-size: .95rem !important;
    padding: .45rem .3rem !important;
    border-radius: 6px;
    transition: background .15s;
    cursor: pointer;
}
[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(124,109,255,.12) !important;
}

/* ═══════════════════════════════════════════════════════════════════════════
   KARTLAR
═══════════════════════════════════════════════════════════════════════════ */
.card {
    background: #161622;
    border-radius: 14px;
    padding: 1.1rem 1.2rem;
    margin-bottom: .65rem;
    border-left: 4px solid #7c6dff;
    border-top: 1px solid #1e1e30;
    border-right: 1px solid #1e1e30;
    border-bottom: 1px solid #1e1e30;
    transition: transform .15s, box-shadow .15s;
}
.card:hover { transform: translateY(-1px); box-shadow: 0 4px 20px rgba(0,0,0,.4); }
.card b { color: #f0f0f8; }
.card-green  { border-left-color: #22c55e; }
.card-orange { border-left-color: #f97316; }
.card-red    { border-left-color: #ef4444; }
.card-gray   { border-left-color: #374151; }
.card-yellow { border-left-color: #eab308; }
.card-blue   { border-left-color: #3b82f6; }

/* ═══════════════════════════════════════════════════════════════════════════
   BUTONLAR
═══════════════════════════════════════════════════════════════════════════ */
.stButton > button {
    border-radius: 10px !important;
    font-weight: 600 !important;
    transition: all .15s !important;
    border: none !important;
    min-height: 44px !important;
}
.stButton > button:hover { transform: translateY(-1px); box-shadow: 0 4px 12px rgba(0,0,0,.4) !important; }
.stButton > button[kind="primary"] { background: linear-gradient(135deg,#7c6dff,#a89bff) !important; color: #fff !important; }
.stFormSubmitButton > button { border-radius: 10px !important; font-weight: 600 !important; min-height: 48px !important; }
.stDownloadButton > button { border-radius: 10px !important; min-height: 44px !important; }

/* ═══════════════════════════════════════════════════════════════════════════
   INPUT / SELECT / FORM
═══════════════════════════════════════════════════════════════════════════ */
[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea,
[data-testid="stNumberInput"] input {
    background: #1a1a2e !important;
    border: 1px solid #2a2a40 !important;
    border-radius: 8px !important;
    color: #e2e2ee !important;
    font-size: .95rem !important;
    transition: border-color .15s !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stTextArea"] textarea:focus,
[data-testid="stNumberInput"] input:focus {
    border-color: #7c6dff !important;
    box-shadow: 0 0 0 2px rgba(124,109,255,.2) !important;
}
[data-testid="stSelectbox"] > div > div {
    background: #1a1a2e !important;
    border: 1px solid #2a2a40 !important;
    border-radius: 8px !important;
}

/* ═══════════════════════════════════════════════════════════════════════════
   METRİKLER
═══════════════════════════════════════════════════════════════════════════ */
[data-testid="stMetric"] {
    background: #161622;
    border-radius: 12px;
    padding: .9rem 1rem;
    border: 1px solid #1e1e30;
}
[data-testid="stMetricValue"] { color: #a89bff !important; font-weight: 700 !important; }
[data-testid="stMetricLabel"] { color: #6b7280 !important; font-size: .82rem !important; }
[data-testid="stMetricDelta"]  { font-size: .8rem !important; }

/* ═══════════════════════════════════════════════════════════════════════════
   TAB, EXPANDER, DIVIDER
═══════════════════════════════════════════════════════════════════════════ */
[data-testid="stTabs"] [data-baseweb="tab-list"] {
    background: #161622;
    border-radius: 10px;
    padding: 4px;
    gap: 2px;
    border: 1px solid #1e1e30;
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    flex-wrap: nowrap;
}
[data-testid="stTabs"] [data-baseweb="tab"] {
    border-radius: 8px !important;
    font-size: .88rem !important;
    font-weight: 600 !important;
    white-space: nowrap !important;
    color: #9ca3af !important;
    padding: .45rem .9rem !important;
    min-height: 40px !important;
}
[data-testid="stTabs"] [aria-selected="true"] {
    background: #7c6dff !important;
    color: #fff !important;
}
[data-testid="stExpander"] {
    background: #161622 !important;
    border: 1px solid #1e1e30 !important;
    border-radius: 12px !important;
}
[data-testid="stExpander"] summary {
    font-weight: 600 !important;
    color: #a89bff !important;
    padding: .8rem 1rem !important;
}
hr { border-color: #1e1e30 !important; margin: 1rem 0 !important; }
[data-testid="stCaptionContainer"] { color: #6b7280 !important; font-size: .82rem !important; }

/* ═══════════════════════════════════════════════════════════════════════════
   PROGRESS BAR
═══════════════════════════════════════════════════════════════════════════ */
[data-testid="stProgress"] > div {
    background: #1e1e30 !important;
    border-radius: 999px !important;
    height: 10px !important;
}
[data-testid="stProgress"] > div > div {
    background: linear-gradient(90deg, #7c6dff, #a89bff) !important;
    border-radius: 999px !important;
}
[data-testid="stProgress"] p { font-size: .82rem !important; color: #9ca3af !important; }

/* ═══════════════════════════════════════════════════════════════════════════
   ÖZEL BİLEŞENLER
═══════════════════════════════════════════════════════════════════════════ */
.next-piece {
    background: linear-gradient(135deg,rgba(124,109,255,.15),rgba(124,109,255,.05));
    border: 2px solid #7c6dff;
    border-radius: 14px;
    padding: 1.2rem;
    text-align: center;
    font-size: 1.05rem;
    font-weight: 700;
    color: #a89bff;
    margin-bottom: .8rem;
    animation: pulse-border 2s infinite;
}
@keyframes pulse-border {
    0%,100% { border-color: #7c6dff; }
    50% { border-color: #a89bff; }
}
.badge {
    display: inline-block;
    padding: .22em .75em;
    border-radius: 999px;
    font-size: .76rem;
    font-weight: 700;
    letter-spacing: .01em;
}
.badge-green  { background:#14532d; color:#86efac; }
.badge-orange { background:#431407; color:#fdba74; }
.badge-red    { background:#450a0a; color:#fca5a5; }
.badge-purple { background:#2e1065; color:#c4b5fd; }
.badge-gray   { background:#1f2937; color:#9ca3af; }
.badge-yellow { background:#422006; color:#fde68a; }
.badge-blue   { background:#172554; color:#93c5fd; }
.sec {
    font-size: 1rem;
    font-weight: 700;
    color: #a89bff;
    border-bottom: 2px solid #1e1e30;
    padding-bottom: .35rem;
    margin-bottom: 1rem;
    letter-spacing: .01em;
}
.sla-bar-wrap { background: #1e1e2e; border-radius: 999px; height: 8px; overflow: hidden; margin-top: .4rem; }
.sla-bar      { height: 8px; border-radius: 999px; transition: width .4s ease; }
.vbox {
    background: #161622;
    border-radius: 12px;
    padding: 1rem .8rem;
    border: 1px solid #1e1e30;
    text-align: center;
    transition: transform .15s;
}
.vbox:hover { transform: translateY(-2px); }

/* Uyarı/bilgi bannerları */
[data-testid="stAlert"] {
    border-radius: 10px !important;
    border-left-width: 4px !important;
    font-size: .92rem !important;
}

/* Dataframe */
[data-testid="stDataFrame"] {
    border-radius: 10px !important;
    overflow: hidden !important;
    border: 1px solid #1e1e30 !important;
}

/* Multiselect tag */
[data-baseweb="tag"] { background: #2e1065 !important; border-radius: 6px !important; }

/* Sayfa başlığı */
h1 { font-weight: 800 !important; letter-spacing: -.02em !important; color: #f0f0f8 !important; }
h2 { font-weight: 700 !important; color: #e2e2ee !important; }
h3 { font-weight: 600 !important; color: #d4d4e8 !important; }

/* ═══════════════════════════════════════════════════════════════════════════
   MOBİL — 768px ve altı (telefon)
═══════════════════════════════════════════════════════════════════════════ */
@media (max-width: 768px) {
    .main .block-container {
        padding: .6rem .6rem 4rem !important;
        max-width: 100% !important;
    }

    /* Tüm kolonlar tek sıra */
    [data-testid="column"] {
        width: 100% !important;
        min-width: 100% !important;
        flex: 1 1 100% !important;
    }
    [data-testid="stHorizontalBlock"] {
        flex-wrap: wrap !important;
        gap: .4rem !important;
    }

    /* Başlıklar */
    h1 { font-size: 1.4rem !important; margin-bottom: .3rem !important; }
    h2 { font-size: 1.2rem !important; }

    /* Butonlar büyük dokunma hedefi */
    .stButton > button,
    .stFormSubmitButton > button,
    .stDownloadButton > button {
        width: 100% !important;
        min-height: 52px !important;
        font-size: 1rem !important;
    }

    /* Kartlar */
    .card { padding: .85rem .9rem !important; font-size: .9rem !important; border-radius: 10px !important; }

    /* Input büyüt */
    [data-testid="stTextInput"] input,
    [data-testid="stNumberInput"] input,
    [data-testid="stTextArea"] textarea {
        font-size: 16px !important;  /* iOS auto-zoom engelle */
        min-height: 48px !important;
    }
    [data-testid="stSelectbox"] > div > div { min-height: 48px !important; font-size: 16px !important; }

    /* Metrik küçült */
    [data-testid="stMetric"] { padding: .6rem .7rem !important; }
    [data-testid="stMetricValue"] { font-size: 1.25rem !important; }

    /* Tabs scroll edilebilir */
    [data-testid="stTabs"] [data-baseweb="tab-list"] {
        padding: 3px !important;
        overflow-x: auto !important;
        scrollbar-width: none !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab-list"]::-webkit-scrollbar { display: none; }
    [data-testid="stTabs"] [data-baseweb="tab"] {
        font-size: .8rem !important;
        padding: .4rem .7rem !important;
    }

    /* Progress text */
    [data-testid="stProgress"] p { font-size: .78rem !important; }

    /* Expander */
    [data-testid="stExpander"] summary { font-size: .9rem !important; padding: .7rem .85rem !important; }

    /* Badge küçük */
    .badge { font-size: .72rem !important; padding: .18em .55em !important; }

    /* Next piece */
    .next-piece { font-size: .95rem !important; padding: 1rem !important; }

    /* Sidebar radio büyüt */
    [data-testid="stSidebar"] .stRadio label { font-size: 1.05rem !important; padding: .55rem .4rem !important; }

    /* Checkbox & toggle büyük */
    .stCheckbox label, .stRadio label { font-size: .95rem !important; line-height: 2.2 !important; }

    /* Dataframe scroll */
    [data-testid="stDataFrame"] { overflow-x: auto !important; -webkit-overflow-scrolling: touch !important; }

    /* Divider ince */
    hr { margin: .7rem 0 !important; }

    /* Alert font */
    [data-testid="stAlert"] { font-size: .88rem !important; padding: .7rem .9rem !important; }
}

/* ═══════════════════════════════════════════════════════════════════════════
   KÜÇÜK TELEFON — 400px ve altı
═══════════════════════════════════════════════════════════════════════════ */
@media (max-width: 400px) {
    .main .block-container { padding: .4rem .4rem 4rem !important; }
    h1 { font-size: 1.2rem !important; }
    .card { padding: .7rem .75rem !important; font-size: .85rem !important; border-radius: 8px !important; }
    [data-testid="stMetricValue"] { font-size: 1.1rem !important; }
    .next-piece { font-size: .88rem !important; padding: .85rem !important; }
}

/* ═══════════════════════════════════════════════════════════════════════════
   TABLET — 769-1024px
═══════════════════════════════════════════════════════════════════════════ */
@media (min-width: 769px) and (max-width: 1024px) {
    .main .block-container { padding: 1rem 1rem 2rem !important; max-width: 100% !important; }
    [data-testid="stHorizontalBlock"]:has([data-testid="column"]:nth-child(5)) [data-testid="column"] {
        min-width: 45% !important;
    }
}

/* ═══════════════════════════════════════════════════════════════════════════
   DOKUNMATIK EKRAN
═══════════════════════════════════════════════════════════════════════════ */
@media (hover: none) and (pointer: coarse) {
    .stButton > button,
    .stFormSubmitButton > button,
    .stDownloadButton > button { min-height: 52px !important; }
    .card { transition: none !important; }
    .card:hover { transform: none !important; }
    .stCheckbox label, .stRadio label { line-height: 2.4 !important; }
}
</style>""", unsafe_allow_html=True)

# ── Session init ──────────────────────────────────────────────────────────────
for k,v in [("user",None),("show_qr",False),("scan_mode",False)]:
    if k not in st.session_state: st.session_state[k]=v

# ══════════════════════════════════════════════════════════════════════════════
#  MÜŞTERİ TAKİP SAYFASI (token ile erişim, login gerektirmez)
# ══════════════════════════════════════════════════════════════════════════════
params = st.query_params
if "track" in params:
    token = params["track"]
    tok_row = q1("SELECT * FROM customer_tokens WHERE token=?", (token,))
    if not tok_row:
        st.error("❌ Geçersiz takip kodu.")
        st.stop()
    # Mark as viewed
    if not tok_row.get("viewed_at"):
        ex("UPDATE customer_tokens SET viewed_at=datetime('now','localtime') WHERE token=?",(token,))
    order = q1("""
        SELECT o.*, c.name collection FROM orders o
        JOIN collections c ON o.collection_id=c.id WHERE o.id=?
    """, (tok_row["order_id"],))
    if not order:
        st.error("Sipariş bulunamadı."); st.stop()

    st.markdown("""
    <style>
    .track-header{text-align:center;padding:1.5rem 1rem .5rem}
    .track-status{background:#1a1a26;border-radius:16px;padding:1.5rem;
                  text-align:center;margin:1rem 0;border:2px solid #7c6dff}
    .track-piece{background:#1a1a26;border-radius:10px;padding:.8rem 1rem;
                 margin-bottom:.4rem;display:flex;align-items:center;gap:.8rem}
    .track-piece-done{opacity:.55}
    </style>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class='track-header'>
      <div style='font-size:2rem'>🪑</div>
      <h2 style='color:#a89bff;margin:.3rem 0'>FurnitureFlow</h2>
      <p style='color:#9ca3af;font-size:.9rem'>Sipariş Takibi</p>
    </div>
    """, unsafe_allow_html=True)

    STATUS_ICONS = {"pending":"⏳ Hazırlanıyor","completed":"✅ Teslimata Hazır"}
    PRIORITY_LABELS = {"urgent":"🚨 ACİL","high":"🔴 Yüksek","normal":"🟡 Normal","low":"🟢 Düşük"}
    items = q("""
        SELECT oi.is_checked, oi.item_status, p.name, p.loading_order
        FROM order_items oi JOIN pieces p ON oi.piece_id=p.id
        WHERE oi.order_id=? ORDER BY p.loading_order
    """, (order["id"],))
    checked = sum(1 for i in items if i["is_checked"])
    total   = len(items)
    pct     = checked/total if total else 0
    pct_int = int(pct*100)

    status_color = "#22c55e" if order["status"]=="completed" else "#f97316"
    status_text  = STATUS_ICONS.get(order["status"],"⏳")

    st.markdown(f"""
    <div class='track-status'>
      <div style='font-size:2.5rem'>{"✅" if order["status"]=="completed" else "⏳"}</div>
      <div style='font-size:1.3rem;font-weight:700;color:{status_color};margin:.4rem 0'>{status_text}</div>
      <div style='font-size:.9rem;color:#9ca3af'>Sipariş No: <b style='color:#e8e8f0'>#{order["order_number"]}</b></div>
      <div style='font-size:.9rem;color:#9ca3af'>Koleksiyon: <b style='color:#e8e8f0'>{order["collection"]}</b></div>
    </div>
    """, unsafe_allow_html=True)

    # Big progress circle simulation with progress bar
    st.markdown(f"""
    <div style='background:#1a1a26;border-radius:12px;padding:1.2rem;margin:.5rem 0'>
      <div style='display:flex;justify-content:space-between;margin-bottom:.5rem'>
        <span style='color:#9ca3af;font-size:.9rem'>Hazırlık Durumu</span>
        <span style='color:#a89bff;font-weight:700;font-size:1.1rem'>%{pct_int}</span>
      </div>
      <div style='background:#2a2a3a;border-radius:8px;height:12px;overflow:hidden'>
        <div style='height:12px;border-radius:8px;width:{pct_int}%;
                    background:{"#22c55e" if pct>=1 else "#7c6dff"};transition:width .3s'></div>
      </div>
      <div style='text-align:center;margin-top:.5rem;color:#9ca3af;font-size:.85rem'>
        {checked} / {total} parça hazır
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Info grid
    st.markdown(f"""
    <div style='display:grid;grid-template-columns:1fr 1fr;gap:.5rem;margin:.5rem 0'>
      <div style='background:#1a1a26;border-radius:10px;padding:.8rem;text-align:center'>
        <div style='color:#9ca3af;font-size:.78rem'>Öncelik</div>
        <div style='font-weight:700;margin-top:.2rem'>{PRIORITY_LABELS.get(order["priority"],"Normal")}</div>
      </div>
      <div style='background:#1a1a26;border-radius:10px;padding:.8rem;text-align:center'>
        <div style='color:#9ca3af;font-size:.78rem'>Sipariş Tarihi</div>
        <div style='font-weight:700;font-size:.85rem;margin-top:.2rem'>{(order["created_at"] or "")[:10]}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    if order["completed_at"]:
        st.markdown(f"""
        <div style='background:#14532d;border-radius:10px;padding:.8rem;text-align:center;margin:.5rem 0'>
          <span style='color:#86efac'>✅ Tamamlanma: {order["completed_at"][:16]}</span>
        </div>
        """, unsafe_allow_html=True)

    st.divider()
    st.markdown("**Parça Listesi**")

    for it in items:
        icon  = "✅" if it["is_checked"] else "⏳"
        s_ico = " ⚠️" if it["item_status"]=="damaged" else " ❌" if it["item_status"]=="missing" else ""
        done_style = "opacity:.5" if it["is_checked"] else ""
        st.markdown(f"""
        <div class='track-piece' style='{done_style}'>
          <span style='font-size:1.3rem'>{icon}</span>
          <span style='flex:1;font-size:.95rem'>{it["name"]}{s_ico}</span>
          <span style='color:#6b7280;font-size:.78rem'>#{it["loading_order"]}</span>
        </div>
        """, unsafe_allow_html=True)

    st.divider()
    st.caption("Bu sayfa size özel oluşturulmuştur. FurnitureFlow · Luxe Life Mobilya")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
#  LOGIN
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.user is None:
    _,col,_ = st.columns([0.5,3,0.5])
    with col:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""
        <div style='text-align:center;margin-bottom:1.8rem'>
          <div style='font-size:3rem;margin-bottom:.5rem'>🪑</div>
          <h1 style='font-size:2rem;font-weight:800;color:#f0f0f8;margin:0;letter-spacing:-.02em'>
            FurnitureFlow
          </h1>
          <p style='color:#6b7280;font-size:.88rem;margin:.4rem 0 0'>
            Luxe Life Mobilya · Lojistik Sistemi
          </p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login"):
            un = st.text_input("Kullanıcı Adı", placeholder="Kullanıcı adınızı girin")
            pw = st.text_input("Şifre", type="password", placeholder="••••••••")
            if st.form_submit_button("🔐 Giriş Yap", use_container_width=True, type="primary"):
                u = login(un, pw)
                if u: st.session_state.user = u; st.rerun()
                else: st.error("❌ Kullanıcı adı veya şifre hatalı.")

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""
        <div style='text-align:center;margin-bottom:.8rem'>
          <span style='color:#6b7280;font-size:.82rem;letter-spacing:.04em;text-transform:uppercase'>
            Demo Hesapları
          </span>
        </div>
        """, unsafe_allow_html=True)

        DEMO_ACCOUNTS = [
            {
                "role":   "🔑 Admin — Tam Yetki",
                "desc":   "Tüm sayfalar · sipariş · analitik · yönetim",
                "user":   "admin",
                "pass":   "admin123",
                "bg":     "#1a1030",
                "border": "#4c1d95",
                "badge":  "#c4b5fd",
                "key":    "qa",
            },
            {
                "role":   "📋 Yönetici",
                "desc":   "Sipariş · rota · analitik · tamamlananlar",
                "user":   "yonetici",
                "pass":   "yonetici123",
                "bg":     "#1a1200",
                "border": "#92400e",
                "badge":  "#fcd34d",
                "key":    "qy",
            },
            {
                "role":   "👷 Depo Personeli",
                "desc":   "Görevlerim · tik atma · stok durumu",
                "user":   "personel",
                "pass":   "personel123",
                "bg":     "#001a0f",
                "border": "#065f46",
                "badge":  "#6ee7b7",
                "key":    "qp",
            },
        ]

        for acc in DEMO_ACCOUNTS:
            st.markdown(f"""
            <div style='background:{acc["bg"]};border-radius:12px;
                        padding:1rem 1.1rem;margin-bottom:.5rem;
                        border:1px solid {acc["border"]}'>
              <div style='font-weight:700;color:{acc["badge"]};
                          font-size:.9rem;margin-bottom:.2rem'>{acc["role"]}</div>
              <div style='font-size:.76rem;color:#9ca3af;
                          margin-bottom:.65rem;line-height:1.4'>{acc["desc"]}</div>
              <div style='display:flex;flex-direction:row;gap:.4rem;
                          align-items:center;flex-wrap:nowrap'>
                <span style='background:rgba(0,0,0,.5);border-radius:6px;
                             padding:.2rem .55rem;font-size:.78rem;color:#e2e2ee;
                             font-family:monospace;border:1px solid rgba(255,255,255,.08);
                             white-space:nowrap'>
                  👤&nbsp;{acc["user"]}
                </span>
                <span style='background:rgba(0,0,0,.5);border-radius:6px;
                             padding:.2rem .55rem;font-size:.78rem;color:#e2e2ee;
                             font-family:monospace;border:1px solid rgba(255,255,255,.08);
                             white-space:nowrap'>
                  🔑&nbsp;{acc["pass"]}
                </span>
              </div>
            </div>
            """, unsafe_allow_html=True)
            # Tek tıkla giriş butonu
            if st.button(f"→ {acc['role'].split('—')[0].strip()} olarak gir",
                         key=acc["key"], use_container_width=True):
                u = login(acc["user"], acc["pass"])
                if u: st.session_state.user = u; st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
    st.stop()

user = st.session_state.user

# ── Role-based default page ────────────────────────────────────────────────────
if "page_initialized" not in st.session_state:
    st.session_state.page_initialized = True
    if user["role"] == "personel":
        st.session_state.default_page = "👷 Görevlerim"
    else:
        st.session_state.default_page = "🏠 Kontrol Paneli"

# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style='padding:.4rem 0 .5rem;display:flex;align-items:center;gap:.55rem'>
      <span style='font-size:1.4rem'>🪑</span>
      <div>
        <div style='font-weight:800;font-size:.98rem;color:#f0f0f8;letter-spacing:-.01em'>FurnitureFlow</div>
        <div style='font-size:.68rem;color:#6b7280'>Luxe Life Mobilya</div>
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    RB = {"admin":("badge-purple","🔑 Admin"),"yonetici":("badge-orange","📋 Yönetici"),"personel":("badge-green","👷 Personel")}
    bc, bl = RB.get(user["role"], ("badge-gray", user["role"]))
    st.markdown(f"""
    <div style='display:flex;align-items:center;gap:.5rem;margin-bottom:.4rem'>
      <div style='width:30px;height:30px;border-radius:50%;background:#2e1065;
                  display:flex;align-items:center;justify-content:center;font-size:.9rem;flex-shrink:0'>
        👤
      </div>
      <div>
        <div style='font-weight:600;font-size:.85rem;color:#e2e2ee;line-height:1.2'>{user['full_name'] or user['username']}</div>
        <span class='badge {bc}' style='font-size:.68rem'>{bl}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    pc  = q1("SELECT COUNT(*) c FROM orders WHERE status='pending'")["c"]
    cc  = q1("SELECT COUNT(*) c FROM orders WHERE status='completed'")["c"]
    unc = unread_count()
    dmg = q1("SELECT COUNT(*) c FROM order_items WHERE item_status IN ('damaged','missing')")["c"]
    active = q("SELECT o.*,c.name collection FROM orders o JOIN collections c ON o.collection_id=c.id WHERE o.status='pending'")
    sla_breach = sum(1 for o in active if sla_status(o)[2] >= 1.0)
    my_orders_count = q1("SELECT COUNT(*) c FROM orders WHERE status='pending' AND assigned_to=?", (user["username"],))["c"]

    alerts = []
    if sla_breach > 0:   alerts.append(f"<span style='color:#fca5a5'>🚨 {sla_breach} SLA</span>")
    if unc > 0:          alerts.append(f"<span style='color:#93c5fd'>🔔 {unc}</span>")
    if dmg > 0 and user["role"] != "personel": alerts.append(f"<span style='color:#fdba74'>⚠️ {dmg}</span>")
    if user["role"] == "personel" and my_orders_count > 0: alerts.append(f"<span style='color:#6ee7b7'>📋 {my_orders_count}</span>")
    if alerts:
        st.markdown(
            "<div style='background:#161622;border-radius:8px;padding:.35rem .65rem;"
            "margin-bottom:.3rem;font-size:.78rem;border:1px solid #1e1e30;line-height:1.9'>"
            + " &nbsp;·&nbsp; ".join(alerts) + "</div>",
            unsafe_allow_html=True
        )

    st.divider()

    if user["role"] == "personel":
        pages = ["👷 Görevlerim", "✅ Aktif Siparişler", "📦 Stok Durumu"]
    elif user["role"] == "yonetici":
        pages = [
            "🏠 Kontrol Paneli", "📝 Yeni Sipariş", "✅ Aktif Siparişler",
            "📋 Tamamlanan Siparişler", "📦 Stok Durumu",
            "🗺️ Rota Optimizasyonu", "📊 Analitik", "🔔 Bildirimler",
        ]
    else:
        pages = [
            "🏠 Kontrol Paneli", "📝 Yeni Sipariş", "✅ Aktif Siparişler",
            "📋 Tamamlanan Siparişler", "📦 Stok Durumu",
            "🗺️ Rota Optimizasyonu", "📊 Analitik", "🔔 Bildirimler",
            "📜 Denetim Kaydı", "⚙️ Admin Paneli",
        ]

    default_idx = 0
    default_pg  = st.session_state.get("default_page", pages[0])
    if default_pg in pages:
        default_idx = pages.index(default_pg)

    page = st.radio("Sayfa", pages, index=default_idx, label_visibility="collapsed")
    st.divider()

    if st.button("🚪 Çıkış Yap", use_container_width=True):
        for k in ["user","page_initialized","default_page","show_qr","scan_mode"]:
            st.session_state.pop(k, None)
        st.rerun()

    st.markdown("""
    <div style='text-align:center;font-size:.65rem;color:#374151;margin-top:.5rem'>
      FurnitureFlow v1.0
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: GÖREVLERİM (Sadece Personel)
# ══════════════════════════════════════════════════════════════════════════════
if page=="👷 Görevlerim":
    name_display = user["full_name"] or user["username"]
    first_name   = name_display.split()[0]
    hour         = datetime.now().hour
    greeting     = "☀️ Günaydın" if hour < 12 else "👋 İyi günler" if hour < 18 else "🌙 İyi akşamlar"
    today_str    = datetime.now().strftime("%-d %B %Y")  # Linux
    try:
        today_str = datetime.now().strftime("%d %B %Y").lstrip("0")
    except:
        today_str = datetime.now().strftime("%d %B %Y")

    # Üst karşılama bandı
    st.markdown(f"""
    <div style='background:linear-gradient(135deg,#1a1030,#0d0820);
                border-radius:14px;padding:1.2rem 1.4rem;margin-bottom:1.2rem;
                border:1px solid #2e1065'>
      <div style='font-size:1.25rem;font-weight:800;color:#f0f0f8'>{greeting}, {first_name}!</div>
      <div style='font-size:.82rem;color:#9ca3af;margin-top:.2rem'>📅 {today_str}</div>
    </div>
    """, unsafe_allow_html=True)

    # Bana atanan aktif siparişler
    my_orders = q("""
        SELECT o.*,c.name collection FROM orders o
        JOIN collections c ON o.collection_id=c.id
        WHERE o.status='pending' AND o.assigned_to=?
        ORDER BY
            CASE o.priority WHEN 'urgent' THEN 0 WHEN 'high' THEN 1
                            WHEN 'normal' THEN 2 ELSE 3 END, o.id ASC
    """, (user["username"],))

    if not my_orders:
        st.markdown("""
        <div style='text-align:center;padding:3.5rem 1rem;background:#161622;
                    border-radius:16px;border:2px dashed #1e1e30;margin:1rem 0'>
          <div style='font-size:3.5rem;margin-bottom:.5rem'>✅</div>
          <div style='font-size:1.2rem;font-weight:700;color:#86efac;margin-bottom:.5rem'>
            Tüm görevler tamamlandı!
          </div>
          <div style='color:#6b7280;font-size:.88rem'>
            Şu an sana atanmış aktif sipariş yok.<br>
            Yöneticin yeni bir görev atadığında burada görünecek.
          </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        PR_COLOR = {"urgent":"#ef4444","high":"#f97316","normal":"#7c6dff","low":"#22c55e"}
        PR_LBL   = {"urgent":"🚨 ACİL","high":"🔴 Yüksek","normal":"🟡 Normal","low":"🟢 Düşük"}
        PR_BG    = {"urgent":"#450a0a","high":"#431407","normal":"#1e1040","low":"#14532d"}

        total_tasks   = len(my_orders)
        urgent_tasks  = sum(1 for o in my_orders if o.get("priority") == "urgent")

        # Mini özet
        c1, c2 = st.columns(2)
        c1.metric("📋 Aktif Görev", total_tasks)
        c2.metric("🚨 Acil", urgent_tasks)
        st.markdown("<br>", unsafe_allow_html=True)

        for o in my_orders:
            items_o   = q("SELECT is_checked FROM order_items WHERE order_id=?", (o["id"],))
            checked_o = sum(1 for i in items_o if i["is_checked"])
            total_o   = len(items_o)
            pct_o     = checked_o / total_o if total_o else 0
            pct_int   = int(pct_o * 100)
            elapsed, sla, sla_pct, sla_lbl, sla_color = sla_status(o)
            pr        = o.get("priority", "normal")
            pr_color  = PR_COLOR.get(pr, "#7c6dff")
            pr_label  = PR_LBL.get(pr, "Normal")
            started   = o.get("started_at")

            # Acil siparişlerde kırmızı border
            card_border = "#ef4444" if pr == "urgent" else "#f97316" if pr == "high" else "#2e1e50"

            st.markdown(f"""
            <div style='background:#161622;border-radius:14px;padding:1.2rem 1.3rem;
                        margin-bottom:.5rem;border:1.5px solid {card_border};
                        box-shadow:0 2px 12px rgba(0,0,0,.3)'>

              <!-- Üst satır: sipariş no + öncelik -->
              <div style='display:flex;justify-content:space-between;
                          align-items:center;margin-bottom:.4rem'>
                <div style='font-size:1.05rem;font-weight:800;color:#f0f0f8'>
                  #{o["order_number"]}
                </div>
                <span style='background:{PR_BG.get(pr,"#1e1040")};color:{pr_color};
                             font-size:.75rem;font-weight:700;padding:.18em .6em;
                             border-radius:999px;border:1px solid {pr_color}44'>
                  {pr_label}
                </span>
              </div>

              <!-- Koleksiyon + müşteri -->
              <div style='font-size:.85rem;color:#9ca3af;margin-bottom:.7rem;line-height:1.5'>
                🗂 {o["collection"]} &nbsp;·&nbsp; 👤 {o["customer_name"] or "Müşteri belirtilmedi"}
              </div>

              <!-- İlerleme çubuğu -->
              <div style='margin-bottom:.5rem'>
                <div style='display:flex;justify-content:space-between;
                            font-size:.78rem;margin-bottom:.3rem'>
                  <span style='color:#6b7280'>Paletleme İlerlemesi</span>
                  <span style='color:#a89bff;font-weight:700'>
                    {checked_o}/{total_o} parça · %{pct_int}
                  </span>
                </div>
                <div style='background:#1e1e30;border-radius:999px;height:10px;overflow:hidden'>
                  <div style='height:10px;border-radius:999px;width:{pct_int}%;
                              background:{"#22c55e" if pct_o >= 1 else "#7c6dff"};
                              transition:width .4s'></div>
                </div>
              </div>

              <!-- SLA (sadece başlatıldıysa) -->
              {f"<div style='font-size:.78rem;color:{sla_color};margin-top:.3rem'>⏱ SLA: {elapsed:.0f} / {sla} dk — {sla_lbl}</div>" if started else ""}
            </div>
            """, unsafe_allow_html=True)

            # Aksiyon butonu — büyük, mobil dostu
            if not started:
                btn_txt  = "▶️  Hazırlamaya Başla"
                btn_type = "primary"
            elif pct_int == 100:
                btn_txt  = "🚀  Tamamla & Çeki Listesi"
                btn_type = "primary"
            else:
                btn_txt  = f"✅  Devam Et  ({total_o - checked_o} parça kaldı)"
                btn_type = "primary"

            if st.button(btn_txt, key=f"go_{o['id']}", use_container_width=True, type=btn_type):
                if not started:
                    ex("UPDATE orders SET started_at=datetime('now','localtime') WHERE id=?", (o["id"],))
                    audit("START", "order", o["id"], f"#{o['order_number']} başlatıldı")
                st.session_state["default_page"]    = "✅ Aktif Siparişler"
                st.session_state["goto_order_label"] = (
                    f"{'🚨 ' if pr=='urgent' else '🔴 ' if pr=='high' else ''}"
                    f"#{o['order_number']} – {o['collection']} ({o['customer_name'] or '—'})"
                )
                st.rerun()
            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    # ── Bugün tamamladıklarım ────────────────────────────────────────────────
    my_done = q("""
        SELECT o.order_number,c.name collection,o.completed_at,o.started_at
        FROM orders o JOIN collections c ON o.collection_id=c.id
        WHERE o.status='completed' AND o.assigned_to=?
          AND DATE(o.completed_at)=DATE('now','localtime')
        ORDER BY o.completed_at DESC
    """, (user["username"],))

    if my_done:
        st.divider()
        st.markdown(f'<div class="sec">🏆 Bugün Tamamladıkların ({len(my_done)})</div>',
                    unsafe_allow_html=True)
        for d in my_done:
            elapsed_d = 0
            if d["started_at"] and d["completed_at"]:
                s2 = datetime.strptime(d["started_at"], "%Y-%m-%d %H:%M:%S")
                e2 = datetime.strptime(d["completed_at"], "%Y-%m-%d %H:%M:%S")
                elapsed_d = int((e2-s2).total_seconds() // 60)
            st.markdown(f"""
            <div class='card card-green' style='padding:.8rem 1rem'>
              <div style='display:flex;justify-content:space-between;align-items:center'>
                <div>
                  <b>#{d["order_number"]}</b>
                  <span style='color:#9ca3af;font-size:.82rem'> — {d["collection"]}</span>
                </div>
                <span style='color:#86efac;font-size:.82rem;font-weight:600'>⏱ {elapsed_d} dk</span>
              </div>
            </div>
            """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: DASHBOARD (Admin & Yönetici)
# ══════════════════════════════════════════════════════════════════════════════
elif page=="🏠 Kontrol Paneli":
    st.title("🏠 Kontrol Paneli")
    st.caption(f"Hoş geldiniz, {user['full_name'] or user['username']} · {datetime.now().strftime('%d %B %Y')}")

    # ── Özet metrikler ───────────────────────────────────────────────────────
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Aktif Sipariş",  pc,
              delta=f"{sla_breach} SLA aşımı" if sla_breach else None,
              delta_color="inverse")
    c2.metric("Bugün Tamamlanan",
              q1("SELECT COUNT(*) c FROM orders WHERE status='completed' AND DATE(completed_at)=DATE('now','localtime')")["c"])
    c3.metric("⚠️ Sorunlu Parça", dmg)
    c4.metric("🔔 Bildirim",     unc)

    st.divider()

    # ── Hızlı İşlemler ──────────────────────────────────────────────────────
    st.markdown('<div class="sec">⚡ Hızlı İşlemler</div>', unsafe_allow_html=True)

    qa1, qa2, qa3, qa4 = st.columns(4)

    with qa1:
        with st.expander("➕ Hızlı Sipariş", expanded=False):
            colls_q = q("SELECT id,name FROM collections ORDER BY name")
            with st.form("quick_order"):
                q_num  = st.text_input("Sipariş No*",
                                       placeholder=f"ORD-2026-{q1('SELECT COUNT(*)+1 c FROM orders')['c']:03d}")
                q_cust = st.text_input("Müşteri Adı")
                q_coll = st.selectbox("Koleksiyon", [c["name"] for c in colls_q])
                q_pr   = st.selectbox("Öncelik",
                                      ["normal","high","urgent","low"],
                                      format_func=lambda x:{"normal":"🟡 Normal","high":"🔴 Yüksek","urgent":"🚨 ACİL","low":"🟢 Düşük"}[x])
                personels_q = q("SELECT username,full_name FROM users WHERE role='personel' AND is_active=1")
                p_opts_q  = ["Atanmamış"] + [p["full_name"] or p["username"] for p in personels_q]
                p_map_q   = {"Atanmamış":""} | {(p["full_name"] or p["username"]):p["username"] for p in personels_q}
                q_assign  = st.selectbox("Personel", p_opts_q)
                if st.form_submit_button("✅ Oluştur", use_container_width=True, type="primary"):
                    if not q_num.strip():
                        st.error("Sipariş No zorunlu")
                    elif q1("SELECT id FROM orders WHERE order_number=?", (q_num,)):
                        st.error("Bu numara zaten var")
                    else:
                        cid_q = next(c["id"] for c in colls_q if c["name"]==q_coll)
                        sla_q = q1("SELECT sla_minutes FROM sla_config WHERE collection_id=?", (cid_q,))
                        oid_q = ex_id("""INSERT INTO orders
                            (order_number,collection_id,customer_name,created_by,
                             assigned_to,priority,sla_minutes)
                            VALUES(?,?,?,?,?,?,?)""",
                            (q_num,cid_q,q_cust,user["username"],
                             p_map_q[q_assign],q_pr,
                             sla_q["sla_minutes"] if sla_q else 90))
                        for p2 in q("SELECT id FROM pieces WHERE collection_id=? ORDER BY loading_order",(cid_q,)):
                            ex("INSERT INTO order_items (order_id,piece_id) VALUES(?,?)",(oid_q,p2["id"]))
                        get_or_create_token(oid_q)
                        push_notif("new_order",f"Hızlı sipariş: #{q_num}",oid_q)
                        audit("CREATE","order",oid_q,f"#{q_num} hızlı oluşturma")
                        st.success(f"✅ #{q_num} oluşturuldu!")
                        st.rerun()

    with qa2:
        with st.expander("👷 Personel Ata", expanded=False):
            unassigned = q("""SELECT o.id,o.order_number,c.name collection
                              FROM orders o JOIN collections c ON o.collection_id=c.id
                              WHERE o.status='pending' AND (o.assigned_to='' OR o.assigned_to IS NULL)
                              ORDER BY o.id DESC LIMIT 10""")
            if unassigned:
                with st.form("quick_assign"):
                    sel_ua = st.selectbox("Sipariş",
                                          [f"#{o['order_number']} ({o['collection']})" for o in unassigned])
                    ua_id  = unassigned[[f"#{o['order_number']} ({o['collection']})" for o in unassigned].index(sel_ua)]["id"]
                    personels_a = q("SELECT username,full_name FROM users WHERE role='personel' AND is_active=1")
                    sel_pa = st.selectbox("Personel", [p["full_name"] or p["username"] for p in personels_a])
                    pa_un  = next((p["username"] for p in personels_a if (p["full_name"] or p["username"])==sel_pa), "")
                    if st.form_submit_button("✅ Ata", use_container_width=True, type="primary"):
                        ex("UPDATE orders SET assigned_to=? WHERE id=?", (pa_un, ua_id))
                        audit("ASSIGN","order",ua_id,f"→ {sel_pa}")
                        st.success(f"✅ Atandı!"); st.rerun()
            else:
                st.info("Atanmamış sipariş yok")

    with qa3:
        with st.expander("🚨 SLA Kritik", expanded=False):
            urgent_orders = [o for o in sorted(active, key=lambda x: sla_status(x)[2], reverse=True) if sla_status(o)[2]>=0.75]
            if urgent_orders:
                for o in urgent_orders[:5]:
                    el,sl,sla_p,sla_l,sla_c = sla_status(o)
                    bw = min(int(sla_p*100),100)
                    st.markdown(f"""
                    <div style='background:#1a1a26;border-radius:8px;padding:.6rem;margin-bottom:.3rem'>
                      <b style='font-size:.9rem'>#{o["order_number"]}</b>
                      <span style='float:right;color:{sla_c};font-size:.82rem'>{sla_l}</span><br>
                      <span style='color:#9ca3af;font-size:.8rem'>{el:.0f}/{sl} dk</span>
                      <div style='background:#2a2a3a;border-radius:4px;height:6px;margin-top:.3rem'>
                        <div style='height:6px;border-radius:4px;width:{bw}%;background:{sla_c}'></div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.success("SLA durumu normal ✅")

    with qa4:
        with st.expander("📊 Bugün Özet", expanded=False):
            today_stats = q1("""
                SELECT
                  COUNT(*) total,
                  SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) done,
                  SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) pending
                FROM orders WHERE DATE(created_at)=DATE('now','localtime')
            """)
            st.metric("Bugün Gelen", today_stats["total"] or 0)
            st.metric("Bugün Tamamlanan", today_stats["done"] or 0)
            st.metric("Bekliyor", today_stats["pending"] or 0)
            if (today_stats["total"] or 0) > 0:
                rate = int((today_stats["done"] or 0)/(today_stats["total"] or 1)*100)
                st.progress(rate/100, text=f"Tamamlanma: %{rate}")

    st.divider()

    # ── Aktif operasyonlar ───────────────────────────────────────────────────
    col_l, col_r = st.columns([3,2])
    with col_l:
        st.markdown('<div class="sec">🔴 Aktif Operasyonlar</div>', unsafe_allow_html=True)
        if active:
            sorted_active = sorted(active, key=priority_score, reverse=True)
            for o in sorted_active[:6]:
                elapsed,sla,sla_p,lbl,color = sla_status(o)
                items_o   = q("SELECT is_checked FROM order_items WHERE order_id=?", (o["id"],))
                checked_o = sum(1 for i in items_o if i["is_checked"])
                total_o   = len(items_o)
                pr        = o.get("priority","normal")
                pr_badge  = {"urgent":"badge-red","high":"badge-orange","normal":"badge-gray","low":"badge-green"}.get(pr,"badge-gray")
                pr_label  = {"urgent":"🚨 ACİL","high":"🔴 Yüksek","normal":"Normal","low":"Düşük"}.get(pr,"Normal")
                card_cls  = "card-red" if sla_p>=1.0 else "card-orange" if sla_p>=0.85 else "card-green"
                bar_w     = min(int(sla_p*100),100)
                assigned  = o.get("assigned_to") or "—"
                pct_items = int(checked_o/total_o*100) if total_o else 0
                st.markdown(f"""
                <div class='card {card_cls}'>
                  <div style='display:flex;justify-content:space-between;flex-wrap:wrap;gap:.3rem'>
                    <b>#{o["order_number"]}</b>
                    <span class='badge {pr_badge}'>{pr_label}</span>
                  </div>
                  <div style='font-size:.85rem;color:#9ca3af;margin:.2rem 0'>
                    🗂 {o["collection"]} &nbsp;|&nbsp; 👷 {assigned} &nbsp;|&nbsp; 👤 {o["customer_name"] or "—"}
                  </div>
                  <div style='display:flex;justify-content:space-between;font-size:.82rem;margin:.3rem 0'>
                    <span style='color:{color}'>⏱ {elapsed:.0f}/{sla} dk — {lbl}</span>
                    <span style='color:#9ca3af'>{checked_o}/{total_o} parça · %{pct_items}</span>
                  </div>
                  <div class='sla-bar-wrap'>
                    <div class='sla-bar' style='width:{bar_w}%;background:{color}'></div>
                  </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style='text-align:center;padding:2rem;background:#1a1a26;border-radius:12px;border:2px dashed #2a2a3a'>
              <div style='font-size:2rem'>✅</div>
              <p style='color:#9ca3af'>Tüm siparişler tamamlandı!</p>
            </div>
            """, unsafe_allow_html=True)

    with col_r:
        # Personel durumu
        st.markdown('<div class="sec">👷 Personel Durumu</div>', unsafe_allow_html=True)
        personels_d = q("SELECT username,full_name FROM users WHERE role='personel' AND is_active=1")
        for p in personels_d:
            p_active = q1("SELECT COUNT(*) c FROM orders WHERE status='pending' AND assigned_to=?", (p["username"],))["c"]
            p_done_today = q1("SELECT COUNT(*) c FROM orders WHERE status='completed' AND assigned_to=? AND DATE(completed_at)=DATE('now','localtime')", (p["username"],))["c"]
            status_color = "#22c55e" if p_active > 0 else "#4b5563"
            st.markdown(f"""
            <div class='card card-gray' style='padding:.7rem 1rem;margin-bottom:.3rem'>
              <div style='display:flex;justify-content:space-between;align-items:center'>
                <div>
                  <span style='font-size:.95rem;font-weight:600'>{p["full_name"] or p["username"]}</span>
                  <span style='display:inline-block;width:8px;height:8px;border-radius:50%;
                               background:{status_color};margin-left:.5rem'></span>
                </div>
                <div style='text-align:right;font-size:.82rem;color:#9ca3af'>
                  📋 {p_active} aktif<br>✅ bugün {p_done_today}
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

        st.divider()
        st.markdown('<div class="sec">Son Tamamlananlar</div>', unsafe_allow_html=True)
        recent = q("""
            SELECT o.order_number,c.name collection,o.customer_name,
                   o.completed_at,o.started_at,o.assigned_to
            FROM orders o JOIN collections c ON o.collection_id=c.id
            WHERE o.status='completed' ORDER BY o.id DESC LIMIT 5
        """)
        for r in recent:
            elapsed_r = 0
            if r["started_at"] and r["completed_at"]:
                s2=datetime.strptime(r["started_at"],"%Y-%m-%d %H:%M:%S")
                e2=datetime.strptime(r["completed_at"],"%Y-%m-%d %H:%M:%S")
                elapsed_r=int((e2-s2).total_seconds()//60)
            st.markdown(f"""
            <div class='card card-green' style='padding:.6rem 1rem;margin-bottom:.25rem'>
              <b>#{r["order_number"]}</b> — {r["collection"]}<br>
              <span style='font-size:.8rem;color:#9ca3af'>
              👷 {r["assigned_to"] or "—"} &nbsp;|&nbsp; ⏱ {elapsed_r} dk
              </span>
            </div>
            """, unsafe_allow_html=True)
        if not recent:
            st.info("Henüz tamamlanan yok.")

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: YENİ SİPARİŞ
# ══════════════════════════════════════════════════════════════════════════════
elif page=="📝 Yeni Sipariş":
    if not can("new_order"): st.error("⛔ Yetki yok"); st.stop()
    st.title("📝 Yeni Sipariş Oluştur")

    colls   = q("SELECT id,name FROM collections ORDER BY name")
    cnames  = [c["name"] for c in colls]
    cids    = {c["name"]:c["id"] for c in colls}

    col_f, col_p = st.columns([1,1])
    with col_f:
        with st.form("new_order"):
            order_num  = st.text_input("Sipariş Numarası *", placeholder="ORD-2026-001")
            cust_name  = st.text_input("Müşteri Adı")
            cust_email = st.text_input("Müşteri E-posta")
            cust_phone = st.text_input("Müşteri Telefon")
            sel_coll   = st.selectbox("Koleksiyon *", cnames)
            priority   = st.selectbox("Öncelik", ["normal","high","urgent","low"],
                                      format_func=lambda x:{"normal":"🟡 Normal","high":"🔴 Yüksek","urgent":"🚨 ACİL","low":"🟢 Düşük"}[x])
            sla_cfg    = q1("SELECT sla_minutes FROM sla_config WHERE collection_id=?", (cids.get(sel_coll,1),))
            sla_val    = st.number_input("SLA Hedefi (dakika)", min_value=15, max_value=480,
                                         value=sla_cfg["sla_minutes"] if sla_cfg else 90)
            personels  = q("SELECT username,full_name FROM users WHERE role='personel' AND is_active=1")
            p_opts     = ["Atanmamış"] + [f"{p['full_name'] or p['username']}" for p in personels]
            p_map      = {"Atanmamış":""} | {(p['full_name'] or p['username']):p['username'] for p in personels}
            assigned   = st.selectbox("Personel Ata", p_opts)
            order_note = st.text_area("Not", height=68)
            submitted  = st.form_submit_button("✅ Sipariş Oluştur", use_container_width=True, type="primary")

        if submitted:
            if not order_num.strip():
                st.error("Sipariş numarası zorunludur.")
            elif q1("SELECT id FROM orders WHERE order_number=?", (order_num,)):
                st.warning(f"'{order_num}' zaten mevcut.")
            else:
                cid = cids[sel_coll]
                oid = ex_id("""
                    INSERT INTO orders
                    (order_number,collection_id,customer_name,customer_email,
                     customer_phone,created_by,assigned_to,priority,sla_minutes,notes)
                    VALUES(?,?,?,?,?,?,?,?,?,?)
                """, (order_num,cid,cust_name,cust_email,cust_phone,
                      user["username"],p_map[assigned],priority,sla_val,order_note))
                pieces_q = q("SELECT id FROM pieces WHERE collection_id=? ORDER BY loading_order",(cid,))
                for p2 in pieces_q:
                    ex("INSERT INTO order_items (order_id,piece_id) VALUES(?,?)",(oid,p2["id"]))
                token   = get_or_create_token(oid)
                app_url = q1("SELECT value FROM app_settings WHERE key='app_url'")
                base    = app_url["value"] if app_url else "http://localhost:8501"
                track_url = f"{base}?track={token}"
                push_notif("new_order", f"Yeni sipariş: #{order_num} ({sel_coll}) — Öncelik: {priority}", oid)
                audit("CREATE", "order", oid, f"#{order_num} | {sel_coll} | {priority} | Atanan: {p_map[assigned]}")
                st.success(f"✅ **{order_num}** oluşturuldu! {len(pieces_q)} parça eklendi.")
                st.code(track_url, language=None)
                st.caption("👆 Bu linki müşteriye gönderin — kendi siparişini takip edebilir.")
                qr_bytes = make_qr(track_url)
                c1,c2 = st.columns(2)
                c1.image(qr_bytes, caption="Müşteri Takip QR", width=180)
                c2.image(make_qr(order_num), caption="Depo Sipariş QR", width=180)
                c1.download_button("⬇️ Takip QR", qr_bytes, f"track_qr_{order_num}.png", "image/png", key="tqr")
                # send tracking email
                if cust_email and cust_email.strip():
                    body = f"""
                    <h2>Merhaba {cust_name or 'Sayın Müşterimiz'},</h2>
                    <p>Siparişiniz sisteme alındı. Durumunu takip etmek için:</p>
                    <p><a href="{track_url}" style="color:#7c6dff;font-size:1.2em">{track_url}</a></p>
                    <p>Sipariş No: <b>{order_num}</b></p>
                    <br><p>Luxe Life Mobilya</p>
                    """
                    ok = send_email(cust_email, f"Siparişiniz Alındı – #{order_num}", body)
                    if ok: st.info("📧 Takip linki müşteriye e-posta ile gönderildi.")

    with col_p:
        st.markdown('<div class="sec">Koleksiyon Önizlemesi</div>',unsafe_allow_html=True)
        prev_coll = st.selectbox("Koleksiyon", cnames, key="prev_coll")
        prev_pcs  = q("""SELECT name,width_cm,depth_cm,height_cm,loading_order
                         FROM pieces WHERE collection_id=(SELECT id FROM collections WHERE name=?)
                         ORDER BY loading_order""", (prev_coll,))
        if prev_pcs:
            tv2 = sum((p["width_cm"]*p["depth_cm"]*p["height_cm"])/1_000_000 for p in prev_pcs)
            sla2= q1("SELECT sla_minutes FROM sla_config WHERE collection_id=(SELECT id FROM collections WHERE name=?)",(prev_coll,))
            c1,c2 = st.columns(2)
            c1.metric("Toplam Hacim", f"{tv2:.3f} m³")
            c2.metric("SLA Hedefi",   f"{sla2['sla_minutes'] if sla2 else 90} dk")
            # Stock warnings
            stock_warn = q("""
                SELECT p.name,ps.status FROM pieces p
                JOIN piece_stock ps ON ps.piece_id=p.id
                JOIN collections c ON p.collection_id=c.id
                WHERE c.name=? AND ps.status!='available'
            """, (prev_coll,))
            if stock_warn:
                st.warning(f"⚠️ {len(stock_warn)} parça stokta yok!")
                for sw in stock_warn:
                    SLBL={"in_production":"Üretimde","at_supplier":"Tedarikçide","waiting":"Bekliyor"}
                    st.caption(f"  • {sw['name']} → {SLBL.get(sw['status'],sw['status'])}")
            for pp in prev_pcs:
                vol=(pp["width_cm"]*pp["depth_cm"]*pp["height_cm"])/1_000_000
                st.markdown(f"""
                <div class='card card-gray' style='padding:.5rem 1rem;margin-bottom:.25rem'>
                  <b>#{pp["loading_order"]} · {pp["name"]}</b><br>
                  <span style='font-size:.82rem;color:#9ca3af'>
                  {pp["width_cm"]}×{pp["depth_cm"]}×{pp["height_cm"]} cm | {vol:.4f} m³
                  </span>
                </div>
                """,unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: AKTİF SİPARİŞLER
# ══════════════════════════════════════════════════════════════════════════════
elif page=="✅ Aktif Siparişler":
    st.title("✅ Aktif Sipariş Takibi")

    active_orders = q("""
        SELECT o.*,c.name collection FROM orders o
        JOIN collections c ON o.collection_id=c.id
        WHERE o.status='pending' ORDER BY o.id DESC
    """)
    if user["role"]=="personel":
        mine = [o for o in active_orders if o.get("assigned_to")==user["username"]]
        active_orders = mine

    if not active_orders:
        msg = "Sana atanmış bekleyen sipariş yok." if user["role"]=="personel" else "Bekleyen sipariş yok."
        st.markdown(f"""
        <div style='text-align:center;padding:3rem 1rem;background:#161622;
                    border-radius:16px;border:2px dashed #1e1e30;margin-top:2rem'>
          <div style='font-size:3rem'>📭</div>
          <div style='font-size:1.1rem;font-weight:700;color:#9ca3af;margin:.5rem 0'>{msg}</div>
          <div style='color:#6b7280;font-size:.88rem'>
          {"👷 Görevlerim sayfasına dön." if user["role"]=="personel" else "Yeni sipariş oluşturulduğunda burada görünür."}
          </div>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    sorted_ao = sorted(active_orders, key=priority_score, reverse=True)
    order_labels = {
        f"{'🚨 ' if o.get('priority')=='urgent' else '🔴 ' if o.get('priority')=='high' else ''}#{o['order_number']} – {o['collection']} ({o['customer_name'] or '—'})": o
        for o in sorted_ao
    }

    goto_label = st.session_state.pop("goto_order_label", None)
    default_ao_idx = 0
    if goto_label:
        for i, lbl in enumerate(order_labels.keys()):
            if goto_label.split("–")[0].strip().lstrip("🚨🔴 ") in lbl:
                default_ao_idx = i; break

    # Sipariş seçici + QR butonu
    col_sel, col_qr = st.columns([5, 1])
    with col_sel:
        sel_lbl = st.selectbox("Sipariş", list(order_labels.keys()),
                               index=default_ao_idx, label_visibility="collapsed")
    sel_o = order_labels[sel_lbl]
    oid   = sel_o["id"]
    with col_qr:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("📷 QR", use_container_width=True):
            st.session_state.show_qr = not st.session_state.show_qr

    if st.session_state.show_qr:
        _,qc,_ = st.columns([1,1,1])
        with qc:
            token2   = get_or_create_token(oid)
            app_url2 = q1("SELECT value FROM app_settings WHERE key='app_url'")
            base2    = app_url2["value"] if app_url2 else "http://localhost:8501"
            track2   = f"{base2}?track={token2}"
            st.image(make_qr(track2), caption="Müşteri Takip", width=160)
            st.image(make_qr(sel_o["order_number"]), caption="Depo Sipariş", width=160)

    # ── Sipariş Bilgi Kartı (okunabilir, büyük) ──────────────────────────────
    PR_LBL   = {"urgent":"🚨 ACİL","high":"🔴 Yüksek","normal":"🟡 Normal","low":"🟢 Düşük"}
    PR_COLOR = {"urgent":"#ef4444","high":"#f97316","normal":"#7c6dff","low":"#22c55e"}
    pr       = sel_o.get("priority","normal")
    pr_color = PR_COLOR.get(pr,"#7c6dff")

    st.markdown(f"""
    <div style='background:#161622;border-radius:14px;padding:1.1rem 1.3rem;
                margin:0.8rem 0;border:1.5px solid {pr_color}55'>
      <div style='display:flex;flex-wrap:wrap;gap:.5rem 1.5rem;align-items:center'>
        <div style='font-size:1rem;font-weight:700;color:{pr_color}'>
          {PR_LBL.get(pr,"Normal")}
        </div>
        <div style='font-size:.95rem;color:#e2e2ee'>
          👷 <b>{sel_o.get("assigned_to") or "Atanmamış"}</b>
        </div>
        <div style='font-size:.95rem;color:#e2e2ee'>
          👤 <b>{sel_o.get("customer_name") or "—"}</b>
        </div>
        <div style='font-size:.88rem;color:#9ca3af'>
          🗂 {sel_o.get("collection","")}
        </div>
      </div>
      {f"<div style='margin-top:.5rem;font-size:.88rem;color:#a89bff'>📝 {sel_o['notes']}</div>" if sel_o.get("notes") else ""}
    </div>
    """, unsafe_allow_html=True)

    # ── Başlat butonu veya SLA çubuğu ───────────────────────────────────────
    if not sel_o.get("started_at"):
        st.button("▶️ Hazırlamaya Başla", type="primary", use_container_width=True,
                  key="start_btn")
        if st.session_state.get("start_btn"):
            ex("UPDATE orders SET started_at=datetime('now','localtime') WHERE id=?", (oid,))
            audit("START","order",oid,f"#{sel_o['order_number']} başlatıldı")
            st.rerun()
        # Güvenli tekrar: butonu ayrı bir if bloğuyla yakala
        if q1("SELECT started_at FROM orders WHERE id=?",(oid,)).get("started_at"):
            st.rerun()
    else:
        elapsed,sla,pct,lbl,color = sla_status(sel_o)
        bar_w = min(int(pct*100),100)

        # Okunabilir SLA bandı — büyük yazı
        st.markdown(f"""
        <div style='background:#161622;border-radius:12px;padding:1rem 1.2rem;
                    border:1px solid #1e1e30;margin:.3rem 0'>
          <div style='display:flex;justify-content:space-between;
                      align-items:center;margin-bottom:.6rem;flex-wrap:wrap;gap:.3rem'>
            <span style='font-size:1rem;font-weight:700;color:#e2e2ee'>
              ⏱ SLA Durumu
            </span>
            <span style='font-size:1rem;font-weight:800;color:{color}'>
              {elapsed:.0f} / {sla} dk — {lbl}
            </span>
          </div>
          <div style='background:#1e1e30;border-radius:999px;height:12px;overflow:hidden'>
            <div style='height:12px;border-radius:999px;width:{bar_w}%;
                        background:{color};transition:width .4s'></div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        col_ref, _ = st.columns([2,5])
        with col_ref:
            auto_refresh = st.toggle("🔄 Canlı Yenileme", value=False, key="auto_refresh",
                                     help="Her 15 saniyede otomatik yenilenir")
        if auto_refresh:
            st.caption(f"🕐 Son yenileme: {datetime.now().strftime('%H:%M:%S')}")
            time.sleep(15)
            st.rerun()

    @st.fragment
    def checklist(oid):
        items = q("""
            SELECT oi.id iid,p.name,p.width_cm,p.depth_cm,p.height_cm,
                   p.loading_order,oi.is_checked,oi.item_status,oi.item_note,oi.checked_at
            FROM order_items oi JOIN pieces p ON oi.piece_id=p.id
            WHERE oi.order_id=? ORDER BY p.loading_order
        """, (oid,))
        checked = sum(1 for i in items if i["is_checked"])
        total   = len(items)
        tv      = sum((i["width_cm"]*i["depth_cm"]*i["height_cm"])/1_000_000 for i in items)
        dmg2    = sum(1 for i in items if i["item_status"] in ("damaged","missing"))
        pct_chk = int(checked/total*100) if total else 0
        remaining = total - checked

        # ── İlerleme özeti — tek geniş kart ─────────────────────────────────
        st.markdown(f"""
        <div style='background:#161622;border-radius:14px;padding:1.1rem 1.3rem;
                    margin-bottom:.8rem;border:1px solid #1e1e30'>

          <!-- İlerleme başlığı -->
          <div style='display:flex;justify-content:space-between;
                      align-items:center;margin-bottom:.6rem;flex-wrap:wrap;gap:.3rem'>
            <span style='font-size:1rem;font-weight:700;color:#e2e2ee'>
              📦 Paletleme Durumu
            </span>
            <span style='font-size:1rem;font-weight:800;
                         color:{"#22c55e" if pct_chk==100 else "#a89bff"}'>
              {checked} / {total} parça &nbsp;·&nbsp; %{pct_chk}
            </span>
          </div>

          <!-- Progress bar -->
          <div style='background:#1e1e30;border-radius:999px;height:14px;
                      overflow:hidden;margin-bottom:.8rem'>
            <div style='height:14px;border-radius:999px;width:{pct_chk}%;
                        background:{"linear-gradient(90deg,#22c55e,#86efac)" if pct_chk==100 else "linear-gradient(90deg,#7c6dff,#a89bff)"};
                        transition:width .4s'></div>
          </div>

          <!-- Alt bilgi satırı -->
          <div style='display:flex;flex-wrap:wrap;gap:.4rem 1.5rem;font-size:.88rem'>
            <span style='color:#9ca3af'>
              ⬜ <b style='color:#e2e2ee'>{remaining}</b> bekliyor
            </span>
            <span style='color:#9ca3af'>
              📦 <b style='color:#e2e2ee'>{tv:.2f} m³</b> toplam hacim
            </span>
            {f"<span style='color:#fdba74'>⚠️ <b>{dmg2}</b> sorunlu parça</span>" if dmg2 > 0 else ""}
            {"<span style='color:#86efac;font-weight:700'>🎉 Tüm parçalar hazır!</span>" if pct_chk == 100 else ""}
          </div>
        </div>
        """, unsafe_allow_html=True)

        # Vehicle capacity
        with st.expander("🚛 Araç Kapasitesi"):
            vs = calc_vehicles(tv)
            vcols = st.columns(len(vs))
            for i,v in enumerate(vs):
                vcols[i].markdown(f"""
                <div class='vbox'>
                  <b>{v["name"]}</b><br>
                  <span style='font-size:1.6rem;color:#a89bff;font-weight:800'>{v["count"]}</span>
                  <span style='color:#9ca3af'> araç</span><br>
                  <span style='font-size:.8rem;color:#9ca3af'>{v["capacity"]} m³/araç</span><br>
                  <span style='color:#86efac'>%{v["fill"]:.0f}</span>
                </div>
                """,unsafe_allow_html=True)

        # Last check / pallet visualization
        with st.expander("📦 Son Kontrol & Palet Yerleşim Önerisi"):
            st.caption("Parçalar büyükten küçüğe yükleme sırası — en ağır/büyük alta konur.")
            items_sorted = sorted(items, key=lambda x: -(x["width_cm"]*x["depth_cm"]*x["height_cm"]))
            for it in items_sorted[:10]:
                vol=(it["width_cm"]*it["depth_cm"]*it["height_cm"])/1_000_000
                bar_len = max(int(vol/tv*30),1) if tv>0 else 1
                icon = "✅" if it["is_checked"] else "⬜"
                st.markdown(f"{icon} **{it['name']}** — {it['width_cm']}×{it['depth_cm']}×{it['height_cm']} cm &nbsp; `{'█'*bar_len}` {vol:.4f} m³")

        st.divider()

        # Barcode scan mode
        col_bm,_ = st.columns([2,5])
        scan_on = col_bm.toggle("🔍 Barkod / QR Tarama Modu", key="scan_mode")
        if scan_on:
            st.info("USB barkod okuyucu veya el terminali bu alana odaklanır. Parça adını veya sipariş numarasını okutun.")
            scan_val = st.text_input("Parça Adı veya Kod", key="scan_input", placeholder="Barkod okutun veya yazın...")
            if scan_val:
                matched = [i for i in items if scan_val.lower() in i["name"].lower() and not i["is_checked"]]
                if matched:
                    it = matched[0]
                    ex("UPDATE order_items SET is_checked=1,checked_at=datetime('now','localtime') WHERE id=?",(it["iid"],))
                    st.success(f"✅ '{it['name']}' işaretlendi!")
                    st.rerun()
                else:
                    st.warning("Eşleşen veya zaten işaretlenmiş parça bulunamadı.")

        st.markdown('<div class="sec">Yükleme Listesi (Uzman Sıralaması)</div>',unsafe_allow_html=True)

        # Next piece highlight
        nxt = next((i for i in items if not i["is_checked"]), None)
        if nxt:
            st.markdown(f"""
            <div class='next-piece'>
              ▶ ŞİMDİ YERLEŞTİR: #{nxt["loading_order"]} · {nxt["name"]}<br>
              <span style='font-size:.85rem;font-weight:400'>
              📐 {nxt["width_cm"]} × {nxt["depth_cm"]} × {nxt["height_cm"]} cm
              </span>
            </div>
            """,unsafe_allow_html=True)

        SOPTS=["normal","damaged","missing"]
        SLBLS={"normal":"✅ Normal","damaged":"⚠️ Hasarlı","missing":"❌ Eksik"}
        SCOLORS={"normal":"#22c55e","damaged":"#f97316","missing":"#ef4444"}

        for it in items:
            vol=(it["width_cm"]*it["depth_cm"]*it["height_cm"])/1_000_000
            done= it["is_checked"]
            s   = it["item_status"] or "normal"

            # Mobil dostu kart yapısı: tüm bilgi tek kartın içinde,
            # buton ve durum seçimi kartın altında yan yana
            fade_style = "opacity:.5;text-decoration:line-through" if done else ""
            s_color    = SCOLORS.get(s,"#22c55e")
            card_cls   = "card-orange" if s=="damaged" else "card-red" if s=="missing" else ("card-gray" if done else "card")

            st.markdown(f"""
            <div class='card {card_cls}' style='margin-bottom:.4rem'>
              <div style='{fade_style}'>
                <span style='font-size:.8rem;color:#6b7280'>#{it["loading_order"]}</span>
                <b style='font-size:1rem'> {it["name"]}</b>
              </div>
              <div style='font-size:.82rem;color:#9ca3af;margin-top:.2rem'>
                📐 {it["width_cm"]}×{it["depth_cm"]}×{it["height_cm"]} cm &nbsp;|&nbsp; 📦 {vol:.4f} m³
                {f"<br>📝 {it['item_note']}" if it.get("item_note") else ""}
              </div>
              <div style='margin-top:.4rem'>
                <span style='font-size:.78rem;color:{s_color};font-weight:600'>
                  {SLBLS[s]}
                </span>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Kontroller: buton ve durum seçimi — mobilde 2 kolon (yeterince geniş)
            btn_col, stat_col = st.columns([1, 2])
            with btn_col:
                if not done:
                    if st.button("⬜ Onayla", key=f"c_{it['iid']}",
                                 use_container_width=True):
                        ex("UPDATE order_items SET is_checked=1,checked_at=datetime('now','localtime') WHERE id=?",(it["iid"],))
                        st.rerun()
                else:
                    if st.button("✅ Geri Al", key=f"u_{it['iid']}",
                                 use_container_width=True):
                        ex("UPDATE order_items SET is_checked=0,checked_at=NULL WHERE id=?",(it["iid"],))
                        st.rerun()
            with stat_col:
                ns=st.selectbox("Durum",SOPTS,format_func=lambda x:SLBLS[x],
                                index=SOPTS.index(s),key=f"s_{it['iid']}",
                                label_visibility="collapsed")
                if ns!=s:
                    ex("UPDATE order_items SET item_status=? WHERE id=?",(ns,it["iid"]))
                    if ns in ("damaged","missing"):
                        push_notif("damage",f"#{q1('SELECT order_number FROM orders WHERE id=?',(oid,))['order_number']}: {it['name']} — {SLBLS[ns]}",oid)
                        audit("DAMAGE", "order_item", it["iid"], f"{it['name']} → {SLBLS[ns]}")
                    st.rerun()

            if ns in ("damaged","missing"):
                nn=st.text_input("Not",value=it["item_note"] or "",key=f"n_{it['iid']}",
                                 placeholder="Hasar/eksik açıklaması...",
                                 label_visibility="collapsed")
                if nn!=(it["item_note"] or ""):
                    ex("UPDATE order_items SET item_note=? WHERE id=?",(nn,it["iid"]))

            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

        st.divider()
        all_ok = checked==total
        if not all_ok:
            st.warning(f"⚠️ Integrity Check: {total-checked} parça henüz onaylanmadı.")
            st.button("🚫 Sevkiyatı Tamamla",disabled=True,use_container_width=True)
        else:
            if dmg2>0: st.warning(f"⚠️ {dmg2} hasarlı/eksik parça var — bunlar Çeki Listesi'ne işlenecek.")
            if st.button("🚀 Sevkiyatı Tamamla & Çeki Listesi Oluştur",use_container_width=True,type="primary"):
                ex("UPDATE orders SET status='completed',completed_at=datetime('now','localtime') WHERE id=?",(oid,))
                push_notif("completed",f"Sipariş tamamlandı: #{sel_o['order_number']}",oid)
                audit("COMPLETE", "order", oid, f"#{sel_o['order_number']} tamamlandı")
                # Send email to customer
                order_full = q1("SELECT o.*,c.name collection FROM orders o JOIN collections c ON o.collection_id=c.id WHERE o.id=?",(oid,))
                if order_full and order_full.get("customer_email"):
                    items_full = q("""SELECT p.name,p.width_cm,p.depth_cm,p.height_cm,
                                       p.loading_order,oi.checked_at,oi.item_status,oi.item_note
                                    FROM order_items oi JOIN pieces p ON oi.piece_id=p.id
                                    WHERE oi.order_id=? ORDER BY p.loading_order""",(oid,))
                    pdf = make_pdf(order_full,items_full)
                    body=f"""<h2>Siparişiniz Hazır!</h2>
                    <p>Sayın {order_full.get('customer_name') or 'Müşterimiz'},</p>
                    <p>Sipariş numaranız <b>{order_full['order_number']}</b> teslimata hazırlanmıştır.</p>
                    <p>Ekte çeki listesi belgenizi bulabilirsiniz.</p><br>
                    <p>Luxe Life Mobilya</p>"""
                    ok=send_email(order_full["customer_email"],f"Siparişiniz Hazır – #{order_full['order_number']}",body,pdf,f"packing_list_{order_full['order_number']}.pdf")
                    if ok: st.info("📧 Müşteriye teslim bildirimi gönderildi.")
                st.success("Sipariş tamamlandı!")
                st.rerun()

    checklist(oid)

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: STOK DURUMU
# ══════════════════════════════════════════════════════════════════════════════
elif page=="📦 Stok Durumu":
    st.title("📦 Parça Stok Durumu")
    st.caption("Her parçanın depo/üretim durumunu buradan yönetin.")

    STOCK_OPTS  = ["available","in_production","at_supplier","waiting"]
    STOCK_LBLS  = {"available":"✅ Depoda","in_production":"🏭 Üretimde",
                   "at_supplier":"🚚 Tedarikçide","waiting":"⏳ Bekliyor"}
    STOCK_COLOR = {"available":"card-green","in_production":"card-blue",
                   "at_supplier":"card-orange","waiting":"card-gray"}

    colls = q("SELECT id,name FROM collections ORDER BY name")
    sel_c = st.selectbox("Koleksiyon", [c["name"] for c in colls])
    cid   = next(c["id"] for c in colls if c["name"]==sel_c)

    # Summary
    summary = q("""
        SELECT ps.status,COUNT(*) cnt FROM piece_stock ps
        JOIN pieces p ON ps.piece_id=p.id
        WHERE p.collection_id=? GROUP BY ps.status
    """, (cid,))
    s_map  = {s["status"]:s["cnt"] for s in summary}
    total_p = q1("SELECT COUNT(*) c FROM pieces WHERE collection_id=?", (cid,))["c"]

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("✅ Depoda",      s_map.get("available",0))
    c2.metric("🏭 Üretimde",    s_map.get("in_production",0))
    c3.metric("🚚 Tedarikçide", s_map.get("at_supplier",0))
    c4.metric("⏳ Bekliyor",    s_map.get("waiting",0))

    # Overall readiness bar
    ready = s_map.get("available",0)
    if total_p > 0:
        st.progress(ready/total_p,
                    text=f"Stok Hazırlığı: {ready}/{total_p} parça depoda (%{int(ready/total_p*100)})")

    st.divider()

    pieces_s = q("""
        SELECT p.id,p.name,p.width_cm,p.depth_cm,p.height_cm,
               p.loading_order,ps.status,ps.note,ps.updated_at
        FROM pieces p
        LEFT JOIN piece_stock ps ON ps.piece_id=p.id
        WHERE p.collection_id=? ORDER BY p.loading_order
    """, (cid,))

    # Filter
    filter_status = st.multiselect("Duruma göre filtrele", STOCK_OPTS,
                                   format_func=lambda x:STOCK_LBLS[x])
    if filter_status:
        pieces_s = [p for p in pieces_s if p["status"] in filter_status]

    st.markdown(f"**{len(pieces_s)} parça gösteriliyor**")
    st.divider()

    if can("admin") or can("new_order"):
        with st.expander("✏️ Toplu Durum Güncelle"):
            st.caption("Bu koleksiyonun görüntülenen tüm parçalarının durumunu tek seferde günceller.")
            with st.form("bulk_stock"):
                bs1, bs2 = st.columns(2)
                new_stat = bs1.selectbox("Yeni Durum", STOCK_OPTS, format_func=lambda x:STOCK_LBLS[x])
                new_note = bs2.text_input("Not (isteğe bağlı)")
                if st.form_submit_button("✅ Tümünü Güncelle", use_container_width=True, type="primary"):
                    for p in pieces_s:
                        ex("UPDATE piece_stock SET status=?,note=?,updated_at=datetime('now','localtime') WHERE piece_id=?",
                           (new_stat,new_note,p["id"]))
                    st.success(f"✅ {len(pieces_s)} parça güncellendi."); st.rerun()

    for p in pieces_s:
        cur_status = p["status"] or "available"
        cc = STOCK_COLOR.get(cur_status,"card-gray")
        # Card with info + controls below (mobile-friendly)
        st.markdown(f"""
        <div class='card {cc}' style='padding:.8rem 1rem;margin-bottom:.3rem'>
          <b>#{p["loading_order"]} · {p["name"]}</b>
          <span style='float:right;font-size:.8rem;color:#9ca3af'>{STOCK_LBLS.get(cur_status,"—")}</span><br>
          <span style='font-size:.82rem;color:#9ca3af'>
          📐 {p["width_cm"]}×{p["depth_cm"]}×{p["height_cm"]} cm
          {f" · 📝 {p['note']}" if p.get("note") else ""}
          </span>
        </div>
        """, unsafe_allow_html=True)

        sc1, sc2 = st.columns([2,3])
        with sc1:
            new_s = st.selectbox("Durum", STOCK_OPTS, format_func=lambda x:STOCK_LBLS[x],
                                 index=STOCK_OPTS.index(cur_status),
                                 key=f"st_{p['id']}", label_visibility="collapsed")
            if new_s != cur_status:
                ex("INSERT OR REPLACE INTO piece_stock (piece_id,status,updated_at) VALUES(?,?,datetime('now','localtime'))",
                   (p["id"],new_s))
                if new_s != "available":
                    push_notif("stock", f"Stok uyarısı: {p['name']} → {STOCK_LBLS[new_s]}")
                    audit("STOCK_UPDATE","piece_stock",p["id"],f"{p['name']} → {STOCK_LBLS[new_s]}")
                st.rerun()
        with sc2:
            nn = st.text_input("Not", value=p.get("note") or "",
                               key=f"sn_{p['id']}", placeholder="Açıklama...",
                               label_visibility="collapsed")
            if nn != (p.get("note") or ""):
                ex("UPDATE piece_stock SET note=?,updated_at=datetime('now','localtime') WHERE piece_id=?",
                   (nn, p["id"]))

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: BİLDİRİMLER
# ══════════════════════════════════════════════════════════════════════════════
elif page=="🔔 Bildirimler":
    if not can("notifications"): st.error("⛔ Yetki yok"); st.stop()
    st.title("🔔 Bildirimler")

    c1,c2 = st.columns([4,1])
    total_notifs = q1("SELECT COUNT(*) c FROM notifications")["c"]
    unread = unread_count()
    c1.caption(f"Toplam {total_notifs} bildirim · {unread} okunmamış")
    if c2.button("✅ Tümünü Oku", use_container_width=True):
        ex("UPDATE notifications SET is_read=1")
        st.rerun()

    # Type filter
    TYPE_ICON  = {"new_order":"📦","completed":"✅","damage":"⚠️","stock":"📊","sla":"🚨","stock_update":"📊"}
    TYPE_LBL   = {"new_order":"Yeni Sipariş","completed":"Tamamlandı","damage":"Hasar",
                  "stock":"Stok","sla":"SLA Uyarısı","stock_update":"Stok Güncelleme"}
    TYPE_CLS   = {"new_order":"card-blue","completed":"card-green","damage":"card-orange",
                  "stock":"card-yellow","sla":"card-red","stock_update":"card-yellow"}

    notifs = q("SELECT * FROM notifications ORDER BY id DESC LIMIT 50")

    if not notifs:
        st.markdown("""
        <div style='text-align:center;padding:2.5rem;background:#1a1a26;border-radius:12px;border:2px dashed #2a2a3a'>
          <div style='font-size:2.5rem'>🔔</div>
          <p style='color:#9ca3af;margin-top:.5rem'>Henüz bildirim yok.</p>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    for n in notifs:
        icon = TYPE_ICON.get(n["type"],"🔔")
        cls  = TYPE_CLS.get(n["type"],"card-gray")
        lbl  = TYPE_LBL.get(n["type"],n["type"])
        fade = "opacity:.55" if n["is_read"] else ""
        read_marker = "" if n["is_read"] else "<span style='display:inline-block;width:8px;height:8px;border-radius:50%;background:#7c6dff;margin-right:6px'></span>"
        st.markdown(f"""
        <div class='card {cls}' style='{fade}padding:.8rem 1.1rem'>
          {read_marker}{icon} <b>{n["message"]}</b><br>
          <span style='font-size:.8rem;color:#9ca3af'>
          🏷 {lbl} &nbsp;|&nbsp; 🕐 {n["created_at"]}
          </span>
        </div>
        """, unsafe_allow_html=True)
        if not n["is_read"]:
            if st.button("✓ Okundu", key=f"nr_{n['id']}", use_container_width=False):
                ex("UPDATE notifications SET is_read=1 WHERE id=?", (n["id"],))
                st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: TAMAMLANAN SİPARİŞLER
# ══════════════════════════════════════════════════════════════════════════════
elif page=="📋 Tamamlanan Siparişler":
    if not can("completed"): st.error("⛔ Yetki yok"); st.stop()
    st.title("📋 Tamamlanan Siparişler")

    f1,f2,f3 = st.columns(3)
    fn  = f1.text_input("🔍 Sipariş No")
    fc  = f2.text_input("🔍 Müşteri")
    fco = f3.selectbox("Koleksiyon",["Tümü"]+[c["name"] for c in q("SELECT name FROM collections ORDER BY name")])

    where,params=["o.status='completed'"],[]
    if fn:  where.append("o.order_number LIKE ?"); params.append(f"%{fn}%")
    if fc:  where.append("o.customer_name LIKE ?"); params.append(f"%{fc}%")
    if fco!="Tümü": where.append("c.name=?"); params.append(fco)
    sql=f"SELECT o.*,c.name collection FROM orders o JOIN collections c ON o.collection_id=c.id WHERE {' AND '.join(where)} ORDER BY o.id DESC"
    done_orders = q(sql,params)

    if not done_orders:
        st.info("Koşula uyan sipariş bulunamadı."); st.stop()
    st.caption(f"{len(done_orders)} sipariş")

    STATUS_ICONS={"normal":"✅","damaged":"⚠️","missing":"❌"}

    for o in done_orders:
        items_d=q("""SELECT p.name,p.width_cm,p.depth_cm,p.height_cm,
                            p.loading_order,oi.checked_at,oi.item_status,oi.item_note
                     FROM order_items oi JOIN pieces p ON oi.piece_id=p.id
                     WHERE oi.order_id=? ORDER BY p.loading_order""",(o["id"],))
        tv3=sum((i["width_cm"]*i["depth_cm"]*i["height_cm"])/1_000_000 for i in items_d)
        dmg3=sum(1 for i in items_d if i["item_status"] in ("damaged","missing"))
        elapsed3,sla3,_,_,_=sla_status(o)
        tag=f" ⚠️{dmg3}sorunlu" if dmg3 else " ✅Sorunsuz"

        with st.expander(f"**#{o['order_number']}** – {o['collection']} | {o['completed_at']}{tag}"):
            c1,c2,c3,c4,c5=st.columns(5)
            c1.metric("Müşteri",  o["customer_name"] or "—")
            c2.metric("Hacim",    f"{tv3:.3f} m³")
            c3.metric("Parça",    len(items_d))
            c4.metric("Sorunlu",  dmg3)
            c5.metric("Hazırlık", f"{elapsed3:.0f} dk")

            for it in items_d:
                vol=(it["width_cm"]*it["depth_cm"]*it["height_cm"])/1_000_000
                icon=STATUS_ICONS.get(it["item_status"] or "normal","✅")
                note=f" — {it['item_note']}" if it["item_note"] else ""
                st.markdown(f"{icon} **#{it['loading_order']} {it['name']}** — {it['width_cm']}×{it['depth_cm']}×{it['height_cm']} cm ({vol:.4f} m³){note}")

            st.divider()
            # Tracking link
            tok=get_or_create_token(o["id"])
            app_url3=q1("SELECT value FROM app_settings WHERE key='app_url'")
            base3=app_url3["value"] if app_url3 else "http://localhost:8501"
            track3=f"{base3}?track={tok}"
            st.code(track3,language=None)
            st.caption("Müşteri bu linkten siparişini takip edebilir.")

            b1,b2,b3=st.columns(3)
            b1.download_button("⬇️ PDF Çeki Listesi",make_pdf(o,items_d),
                               f"packing_list_{o['order_number']}.pdf","application/pdf",
                               use_container_width=True,key=f"pdf_{o['id']}")
            b2.download_button("📊 Excel Çeki Listesi",make_xlsx(o,items_d),
                               f"packing_list_{o['order_number']}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True,key=f"xlsx_{o['id']}")
            if b3.button("📧 Müşteriye Gönder",use_container_width=True,key=f"mail_{o['id']}"):
                if o.get("customer_email"):
                    pdf=make_pdf(o,items_d)
                    body=f"""<h2>Siparişiniz Hazır!</h2><p>Sipariş No: <b>{o['order_number']}</b></p>
                    <p>Ekte çeki listenizi bulabilirsiniz.</p><br><p>Luxe Life Mobilya</p>"""
                    ok=send_email(o["customer_email"],f"Çeki Listesi – #{o['order_number']}",body,pdf,f"packing_list_{o['order_number']}.pdf")
                    if ok: st.success("📧 Gönderildi!")
                else: st.warning("Müşteri e-postası yok.")

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: ANALİTİK
# ══════════════════════════════════════════════════════════════════════════════
elif page=="📊 Analitik":
    if not can("analytics"): st.error("⛔ Yetki yok"); st.stop()
    st.title("📊 Analitik")

    t1,t2,t3,t4,t5 = st.tabs(["📈 Genel","🗂 Koleksiyon","⚠️ Hasar","👷 Personel","🔮 Talep Tahmini"])

    # ── Genel ──
    with t1:
        total_all=pc+cc
        rate=int(cc/total_all*100) if total_all>0 else 0
        avg=q1("SELECT AVG((julianday(completed_at)-julianday(started_at))*24*60) v FROM orders WHERE status='completed' AND started_at IS NOT NULL AND completed_at IS NOT NULL")
        avg_v=avg["v"]
        sla_breach_cnt=q1("SELECT COUNT(*) c FROM orders WHERE status='completed' AND started_at IS NOT NULL AND completed_at IS NOT NULL AND (julianday(completed_at)-julianday(started_at))*24*60 > sla_minutes")["c"]

        c1,c2,c3,c4=st.columns(4)
        c1.metric("Toplam Sipariş",     total_all)
        c2.metric("Tamamlanma Oranı",   f"%{rate}")
        c3.metric("Ort. Hazırlık",      f"{avg_v:.0f} dk" if avg_v else "—")
        c4.metric("SLA Aşımı (tamamlanan)", sla_breach_cnt)

        st.divider()
        st.markdown("**Son 60 Günlük Sipariş Trendi**")
        daily=q("SELECT DATE(created_at) day,COUNT(*) cnt FROM orders WHERE created_at>=DATE('now','-60 days') GROUP BY DATE(created_at) ORDER BY day")
        if daily:
            st.bar_chart(pd.DataFrame(daily).set_index("day").rename(columns={"cnt":"Sipariş"}))
        else: st.info("Yeterli veri yok.")

        st.divider()
        st.markdown("**Öncelik Dağılımı**")
        prio=q("SELECT priority,COUNT(*) cnt FROM orders GROUP BY priority")
        if prio:
            PL2={"urgent":"🚨 ACİL","high":"🔴 Yüksek","normal":"🟡 Normal","low":"🟢 Düşük"}
            df_p=pd.DataFrame(prio)
            df_p["priority"]=df_p["priority"].map(lambda x:PL2.get(x,x))
            st.bar_chart(df_p.set_index("priority").rename(columns={"cnt":"Sipariş"}))

    # ── Koleksiyon ──
    with t2:
        st.markdown("**Koleksiyon Bazında Ortalama Hazırlık Süresi**")
        coll_avg=q("""
            SELECT c.name,ROUND(AVG((julianday(o.completed_at)-julianday(o.started_at))*24*60),1) avg_min,
                   ROUND(AVG(o.sla_minutes),0) avg_sla, COUNT(o.id) cnt
            FROM orders o JOIN collections c ON o.collection_id=c.id
            WHERE o.status='completed' AND o.started_at IS NOT NULL AND o.completed_at IS NOT NULL
            GROUP BY c.id ORDER BY avg_min DESC
        """)
        if coll_avg:
            df_ca=pd.DataFrame(coll_avg).set_index("name")[["avg_min","avg_sla"]]
            df_ca.columns=["Ort. Hazırlık (dk)","Ort. SLA (dk)"]
            st.bar_chart(df_ca)
            st.dataframe(pd.DataFrame(coll_avg),use_container_width=True,hide_index=True)
        else: st.info("Yeterli veri yok.")

        st.divider()
        st.markdown("**Koleksiyon Parça & Hacim Özeti**")
        summary=q("""SELECT c.name Koleksiyon,COUNT(p.id) "Parça",
                     ROUND(SUM(p.width_cm*p.depth_cm*p.height_cm)/1000000.0,3) "Toplam Hacim (m³)"
                     FROM collections c JOIN pieces p ON p.collection_id=c.id
                     GROUP BY c.id ORDER BY 3 DESC""")
        if summary: st.dataframe(pd.DataFrame(summary),use_container_width=True,hide_index=True)

    # ── Hasar ──
    with t3:
        dmg_data=q("""
            SELECT p.name piece,c.name collection,oi.item_status,oi.item_note,o.order_number,o.customer_name
            FROM order_items oi JOIN pieces p ON oi.piece_id=p.id
            JOIN orders o ON oi.order_id=o.id JOIN collections c ON p.collection_id=c.id
            WHERE oi.item_status IN ('damaged','missing') ORDER BY o.id DESC
        """)
        if dmg_data:
            td=sum(1 for d in dmg_data if d["item_status"]=="damaged")
            tm=sum(1 for d in dmg_data if d["item_status"]=="missing")
            c1,c2,c3=st.columns(3); c1.metric("Toplam",len(dmg_data)); c2.metric("⚠️ Hasarlı",td); c3.metric("❌ Eksik",tm)
            st.divider()
            st.markdown("**En Sorunlu Parçalar**")
            top_dmg=q("""SELECT p.name,COUNT(*) cnt FROM order_items oi JOIN pieces p ON oi.piece_id=p.id
                         WHERE oi.item_status IN ('damaged','missing') GROUP BY p.id ORDER BY cnt DESC LIMIT 10""")
            if top_dmg:
                st.bar_chart(pd.DataFrame(top_dmg).set_index("name").rename(columns={"cnt":"Sorun"}))
            st.divider()
            for d in dmg_data:
                cls="card-orange" if d["item_status"]=="damaged" else "card-red"
                icon="⚠️" if d["item_status"]=="damaged" else "❌"
                lbl="Hasarlı" if d["item_status"]=="damaged" else "Eksik"
                note=f" | {d['item_note']}" if d["item_note"] else ""
                st.markdown(f"""
                <div class='card {cls}'>
                  {icon} <b>{d["piece"]}</b> ({d["collection"]}) — {lbl}<br>
                  <span style='color:#9ca3af;font-size:.85rem'>
                  #{d["order_number"]} | {d["customer_name"] or "—"}{note}
                  </span>
                </div>
                """,unsafe_allow_html=True)
        else: st.success("🎉 Hiç hasar/eksik kaydı yok!")

    # ── Personel Performans ──
    with t4:
        st.markdown("**Personel Bazında Sipariş Performansı**")
        perf=q("""
            SELECT assigned_to,
                   COUNT(*) total,
                   SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) completed,
                   ROUND(AVG(CASE WHEN status='completed' AND started_at IS NOT NULL AND completed_at IS NOT NULL
                             THEN (julianday(completed_at)-julianday(started_at))*24*60 END),1) avg_min,
                   SUM(CASE WHEN status='completed' AND started_at IS NOT NULL AND completed_at IS NOT NULL
                            AND (julianday(completed_at)-julianday(started_at))*24*60 > sla_minutes
                            THEN 1 ELSE 0 END) sla_breach
            FROM orders WHERE assigned_to IS NOT NULL AND assigned_to!=''
            GROUP BY assigned_to ORDER BY completed DESC
        """)
        if perf:
            st.dataframe(pd.DataFrame(perf).rename(columns={
                "assigned_to":"Personel","total":"Toplam","completed":"Tamamlanan",
                "avg_min":"Ort. Süre (dk)","sla_breach":"SLA Aşımı"
            }),use_container_width=True,hide_index=True)
            st.divider()
            st.markdown("**Ort. Hazırlık Süresi (dk)**")
            df_perf=pd.DataFrame([p for p in perf if p["avg_min"]]).set_index("assigned_to")[["avg_min"]]
            df_perf.columns=["Ort. Süre (dk)"]
            if not df_perf.empty: st.bar_chart(df_perf)
        else: st.info("Atanmış sipariş verisi yok.")

        st.divider()
        st.markdown("**Günlük Tamamlama Verimliliği**")
        daily_perf=q("""
            SELECT DATE(completed_at) day,assigned_to,COUNT(*) cnt
            FROM orders WHERE status='completed' AND assigned_to!='' AND completed_at>=DATE('now','-30 days')
            GROUP BY DATE(completed_at),assigned_to ORDER BY day
        """)
        if daily_perf:
            df_dp=pd.DataFrame(daily_perf).pivot_table(index="day",columns="assigned_to",values="cnt",fill_value=0)
            st.line_chart(df_dp)
        else: st.info("Yeterli veri yok.")

    # ── Talep Tahmini ──
    with t5:
        st.markdown("**Koleksiyon Bazında Talep Tahmini (Sonraki 30 Gün)**")
        st.caption("Geçmiş 90 günlük veriden hareketli ortalama ile hesaplanır.")

        hist=q("""
            SELECT c.name collection,DATE(o.created_at) day,COUNT(*) cnt
            FROM orders o JOIN collections c ON o.collection_id=c.id
            WHERE o.created_at>=DATE('now','-90 days')
            GROUP BY c.name,DATE(o.created_at) ORDER BY day
        """)
        if hist:
            df_h=pd.DataFrame(hist)
            colls_u=df_h["collection"].unique()
            forecasts=[]
            for col_u in colls_u:
                df_c=df_h[df_h["collection"]==col_u].copy()
                df_c=df_c.set_index("day")["cnt"]
                all_days=pd.date_range(end=pd.Timestamp.now(),periods=90,freq="D").strftime("%Y-%m-%d")
                df_c=df_c.reindex(all_days,fill_value=0)
                avg7  = df_c.tail(7).mean()
                avg30 = df_c.tail(30).mean()
                avg90 = df_c.mean()
                # Weighted forecast
                forecast = round(avg7*0.5 + avg30*0.35 + avg90*0.15, 1) * 30
                forecasts.append({"Koleksiyon":col_u,"Son 7 gün ort/gün":round(avg7,2),
                                  "Son 30 gün ort/gün":round(avg30,2),"30 Günlük Tahmin":round(forecast,0)})
            if forecasts:
                df_f=pd.DataFrame(forecasts).sort_values("30 Günlük Tahmin",ascending=False)
                st.dataframe(df_f,use_container_width=True,hide_index=True)
                st.bar_chart(df_f.set_index("Koleksiyon")[["30 Günlük Tahmin"]])
                st.caption("⚠️ Tahmin, geçmiş veriye dayanır. Kampanya/sezon dönemleri sonucu etkileyebilir.")
        else: st.info("Tahmin için yeterli veri yok. Önce örnek veri üretin.")

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: ROTA OPTİMİZASYONU
# ══════════════════════════════════════════════════════════════════════════════
elif page=="🗺️ Rota Optimizasyonu":
    if not can("new_order"): st.error("⛔ Yetki yok"); st.stop()
    st.title("🗺️ Rota Optimizasyonu")
    st.caption("Aynı gün teslim edilecek siparişler için en kısa rotayı hesaplar.")

    tab_plan, tab_add, tab_history = st.tabs(["📍 Rota Planla", "➕ Teslimat Adresi Ekle", "📋 Geçmiş Planlar"])

    # ── Rota Planla ──────────────────────────────────────────────────────────
    with tab_plan:
        sel_date = st.date_input("Teslimat Tarihi", value=datetime.now().date())
        sel_date_str = sel_date.strftime("%Y-%m-%d")

        # O güne ait teslimatları getir
        day_deliveries = q("""
            SELECT d.*, o.order_number, o.customer_name, o.customer_phone,
                   c.name collection, o.priority
            FROM deliveries d
            JOIN orders o ON d.order_id=o.id
            JOIN collections c ON o.collection_id=c.id
            WHERE d.delivery_date=? AND d.status!='delivered'
            ORDER BY o.priority DESC
        """, (sel_date_str,))

        if not day_deliveries:
            st.info(f"**{sel_date_str}** tarihinde planlanmış teslimat yok.")
            st.caption("'Teslimat Adresi Ekle' sekmesinden tamamlanmış siparişlere adres ekleyebilirsin.")
        else:
            # Koordinatı olmayan teslimatları kontrol et
            no_coords = [d for d in day_deliveries if d["lat"]==0 and d["lng"]==0]
            if no_coords:
                st.warning(f"⚠️ {len(no_coords)} teslimatın koordinatı eksik. İlçe seçimi varsa kullanılacak, yoksa Bursa merkez kabul edilecek.")

            # Koordinat yoksa ilçeye göre ata
            stops = []
            for d in day_deliveries:
                lat, lng = d["lat"], d["lng"]
                if lat==0 and lng==0 and d["district"]:
                    lat, lng = BURSA_DISTRICTS.get(d["district"], (40.1826, 29.0665))
                elif lat==0 and lng==0:
                    lat, lng = 40.1826, 29.0665
                stops.append({
                    "id":           d["id"],
                    "order_id":     d["order_id"],
                    "order_number": d["order_number"],
                    "name":         d["customer_name"] or "—",
                    "phone":        d["customer_phone"] or "—",
                    "address":      f"{d['address']}, {d['district']}, {d['city']}",
                    "district":     d["district"],
                    "collection":   d["collection"],
                    "priority":     d["priority"],
                    "time_window":  d["time_window"],
                    "lat":          lat,
                    "lng":          lng,
                })

            c_veh, c_dep = st.columns(2)
            vehicles_l = q("SELECT name FROM vehicle_types ORDER BY capacity_m3 DESC")
            vehicle_sel = c_veh.selectbox("Araç", [v["name"] for v in vehicles_l])
            depot_district = c_dep.selectbox("Depo Konumu", list(BURSA_DISTRICTS.keys()), index=0)
            depot_lat, depot_lng = BURSA_DISTRICTS[depot_district]

            if st.button("🔄 Rota Hesapla", type="primary", use_container_width=True):
                route, total_km = nearest_neighbor_route(stops, depot_lat, depot_lng)

                st.success(f"✅ **{len(route)} durak** | Tahmini toplam mesafe: **{total_km} km**")
                st.divider()

                # Rota görselleştirme
                st.markdown('<div class="sec">Önerilen Rota Sırası</div>', unsafe_allow_html=True)
                st.markdown(f"""
                <div class='card card-blue' style='padding:.8rem 1.2rem'>
                  🏭 <b>DEPO</b> ({depot_district}) → {'→'.join([f"#{i+1}" for i in range(len(route))])} → 🏭 DEPO<br>
                  <span style='color:#9ca3af;font-size:.85rem'>Toplam: ~{total_km} km | Tahmini süre: ~{int(total_km/40*60)} dk</span>
                </div>
                """, unsafe_allow_html=True)

                for i, stop in enumerate(route, 1):
                    PR_CLS = {"urgent":"card-red","high":"card-orange","normal":"card-green","low":"card-gray"}
                    PR_LBL2 = {"urgent":"🚨 ACİL","high":"🔴 Yüksek","normal":"🟡 Normal","low":"🟢 Düşük"}
                    cls = PR_CLS.get(stop["priority"], "card-gray")
                    st.markdown(f"""
                    <div class='card {cls}'>
                      <b>Durak {i}</b> &nbsp; <span class='badge badge-purple'>+{stop["dist_from_prev"]} km</span><br>
                      👤 <b>{stop["name"]}</b> ({stop["collection"]}) &nbsp; {PR_LBL2.get(stop["priority"],"Normal")}<br>
                      📍 {stop["address"]}<br>
                      <span style='font-size:.82rem;color:#9ca3af'>
                      📞 {stop["phone"]} &nbsp;|&nbsp; 🕐 {stop["time_window"]}
                      </span>
                    </div>
                    """, unsafe_allow_html=True)

                # Save plan
                stop_ids = [str(s["order_id"]) for s in route]
                ex("""INSERT INTO route_plans
                      (plan_date,vehicle_name,stop_order,total_km_est,created_by)
                      VALUES(?,?,?,?,?)""",
                   (sel_date_str, vehicle_sel, ",".join(stop_ids), total_km,
                    user["username"]))
                audit("ROUTE_PLAN", "route_plans", None,
                      f"{sel_date_str} | {vehicle_sel} | {len(route)} durak | {total_km} km")

                # Export as text
                lines = [f"ROTA PLANI — {sel_date_str}", f"Araç: {vehicle_sel}", f"Depo: {depot_district}", ""]
                lines.append(f"DEPO ({depot_district})")
                for i, s in enumerate(route, 1):
                    lines.append(f"  ↓ {s['dist_from_prev']} km")
                    lines.append(f"Durak {i}: {s['name']} | {s['address']} | {s['time_window']}")
                lines.append(f"  ↓ Depoya dönüş")
                lines.append(f"Toplam: ~{total_km} km")
                st.download_button("⬇️ Rota TXT İndir", "\n".join(lines),
                                   f"rota_{sel_date_str}.txt", "text/plain",
                                   use_container_width=True)

    # ── Teslimat Adresi Ekle ─────────────────────────────────────────────────
    with tab_add:
        st.markdown("**Tamamlanmış siparişlere teslimat adresi ekle**")

        # Orders without delivery record
        no_delivery = q("""
            SELECT o.id, o.order_number, o.customer_name, c.name collection
            FROM orders o JOIN collections c ON o.collection_id=c.id
            WHERE o.status='completed'
              AND o.id NOT IN (SELECT order_id FROM deliveries)
            ORDER BY o.id DESC LIMIT 50
        """)

        # Orders with delivery
        has_delivery = q("""
            SELECT o.id, o.order_number, o.customer_name, d.delivery_date,
                   d.address, d.district, d.status
            FROM orders o JOIN deliveries d ON d.order_id=o.id
            JOIN collections c ON o.collection_id=c.id
            ORDER BY d.delivery_date DESC, o.id DESC LIMIT 30
        """)

        col_new, col_exist = st.columns([1, 1])

        with col_new:
            st.markdown("**📍 Yeni Teslimat Adresi**")
            if no_delivery:
                order_map = {f"#{o['order_number']} — {o['customer_name'] or '—'} ({o['collection']})": o
                             for o in no_delivery}
                sel_order_d = st.selectbox("Sipariş", list(order_map.keys()), key="del_order")
                sel_o_d = order_map[sel_order_d]

                with st.form("add_delivery"):
                    address     = st.text_input("Adres", placeholder="Mahalle, sokak, no...")
                    district    = st.selectbox("İlçe", list(BURSA_DISTRICTS.keys()))
                    city        = st.text_input("Şehir", value="Bursa")
                    del_date    = st.date_input("Teslimat Tarihi", value=datetime.now().date())
                    time_window = st.selectbox("Zaman Dilimi",
                                              ["09:00-12:00","12:00-15:00","15:00-18:00","09:00-18:00","Tüm Gün"])
                    # Optional manual coordinates
                    use_coords = st.checkbox("Manuel Koordinat Gir (isteğe bağlı)")
                    if use_coords:
                        cc1, cc2 = st.columns(2)
                        man_lat = cc1.number_input("Enlem", value=40.1826, format="%.4f")
                        man_lng = cc2.number_input("Boylam", value=29.0665, format="%.4f")
                    else:
                        man_lat, man_lng = BURSA_DISTRICTS.get(district, (40.1826, 29.0665))

                    if st.form_submit_button("📍 Adresi Kaydet", use_container_width=True, type="primary"):
                        ex("""INSERT INTO deliveries
                              (order_id,address,district,city,lat,lng,delivery_date,time_window)
                              VALUES(?,?,?,?,?,?,?,?)""",
                           (sel_o_d["id"], address, district, city,
                            man_lat, man_lng, del_date.strftime("%Y-%m-%d"), time_window))
                        audit("ADD_DELIVERY", "deliveries", sel_o_d["id"],
                              f"#{sel_o_d['order_number']} → {district}, {city}")
                        st.success(f"✅ '{sel_o_d['order_number']}' için teslimat adresi eklendi!")
                        st.rerun()
            else:
                st.info("Adresi olmayan tamamlanmış sipariş yok.")

        with col_exist:
            st.markdown("**📋 Mevcut Teslimatlar**")
            if has_delivery:
                DSTATUS = {"pending":"⏳ Bekliyor","in_transit":"🚚 Yolda","delivered":"✅ Teslim"}
                for d in has_delivery:
                    cls = "card-green" if d["status"]=="delivered" else "card-orange" if d["status"]=="in_transit" else "card-gray"
                    st.markdown(f"""
                    <div class='card {cls}' style='padding:.6rem 1.2rem;margin-bottom:.3rem'>
                      <b>#{d["order_number"]}</b> — {d["customer_name"] or "—"}<br>
                      <span style='font-size:.82rem;color:#9ca3af'>
                      📍 {d["district"]} | 📅 {d["delivery_date"]} | {DSTATUS.get(d["status"],"—")}
                      </span>
                    </div>
                    """, unsafe_allow_html=True)

                # Status update
                st.divider()
                st.markdown("**Durum Güncelle**")
                upd_map = {f"#{d['order_number']} ({d['district']})": d for d in has_delivery}
                upd_sel = st.selectbox("Teslimat", list(upd_map.keys()), key="upd_del")
                upd_d   = upd_map[upd_sel]
                new_dstatus = st.selectbox("Yeni Durum",
                                           ["pending","in_transit","delivered"],
                                           format_func=lambda x: DSTATUS[x],
                                           index=["pending","in_transit","delivered"].index(upd_d["status"]),
                                           key="new_dstatus")
                if st.button("💾 Durumu Güncelle", use_container_width=True):
                    ex("UPDATE deliveries SET status=? WHERE order_id=?", (new_dstatus, upd_d["id"]))
                    audit("UPDATE_DELIVERY", "deliveries", upd_d["id"],
                          f"#{upd_d['order_number']} → {DSTATUS[new_dstatus]}")
                    st.success("✅ Güncellendi!"); st.rerun()
            else:
                st.info("Henüz adres eklenmemiş.")

    # ── Geçmiş Planlar ───────────────────────────────────────────────────────
    with tab_history:
        plans = q("SELECT * FROM route_plans ORDER BY id DESC LIMIT 20")
        if plans:
            for p in plans:
                stop_ids = p["stop_order"].split(",") if p["stop_order"] else []
                st.markdown(f"""
                <div class='card card-blue' style='padding:.8rem 1.2rem;margin-bottom:.4rem'>
                  <b>📅 {p["plan_date"]}</b> — {p["vehicle_name"]}<br>
                  🛑 {len(stop_ids)} durak &nbsp;|&nbsp; ~{p["total_km_est"]} km &nbsp;|&nbsp;
                  ⏱ ~{int(p["total_km_est"]/40*60)} dk &nbsp;|&nbsp;
                  👤 {p["created_by"]} &nbsp;|&nbsp; 🕐 {p["created_at"]}
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("Henüz rota planı oluşturulmamış.")

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: DENETİM KAYDI (AUDIT LOG)
# ══════════════════════════════════════════════════════════════════════════════
elif page=="📜 Denetim Kaydı":
    if not can("analytics"): st.error("⛔ Yetki yok"); st.stop()
    st.title("📜 Denetim Kaydı")
    st.caption("Sistemde gerçekleştirilen tüm işlemlerin kronolojik kaydı.")

    # Filters
    f1, f2, f3, f4 = st.columns(4)
    f_user = f1.text_input("👤 Kullanıcı", placeholder="kullanıcı adı...")

    ACTION_TR_MAP = {
        "Tümü":               None,
        "📦 Oluşturuldu":     "CREATE",
        "✅ Tamamlandı":      "COMPLETE",
        "⚠️ Hasar":           "DAMAGE",
        "▶️ Başlatıldı":      "START",
        "👷 Personel Atandı": "ASSIGN",
        "🚚 Teslimat Güncellendi": "UPDATE_DELIVERY",
        "📍 Teslimat Eklendi": "ADD_DELIVERY",
        "🗺️ Rota Planlandı":  "ROUTE_PLAN",
        "🗑️ Silindi":         "DELETE",
        "🔐 Giriş":           "LOGIN",
    }
    ENTITY_TR_MAP = {
        "Tümü":           None,
        "Sipariş":        "order",
        "Sipariş Kalemi": "order_item",
        "Teslimat":       "deliveries",
        "Rota Planı":     "route_plans",
        "Parça":          "pieces",
        "Kullanıcı":      "users",
        "Stok":           "piece_stock",
    }

    sel_action = f2.selectbox("İşlem Türü", list(ACTION_TR_MAP.keys()))
    sel_entity = f3.selectbox("Varlık Türü", list(ENTITY_TR_MAP.keys()))
    f_date     = f4.date_input("Tarihten", value=datetime.now().date()-timedelta(days=7))

    f_action_db = ACTION_TR_MAP[sel_action]
    f_entity_db = ENTITY_TR_MAP[sel_entity]

    where = ["created_at >= ?"]
    params = [f_date.strftime("%Y-%m-%d")]
    if f_user:          where.append("username LIKE ?"); params.append(f"%{f_user}%")
    if f_action_db:     where.append("action=?");        params.append(f_action_db)
    if f_entity_db:     where.append("entity=?");        params.append(f_entity_db)

    logs = q(f"""
        SELECT * FROM audit_log
        WHERE {' AND '.join(where)}
        ORDER BY id DESC LIMIT 200
    """, params)

    st.caption(f"{len(logs)} kayıt bulundu.")
    st.divider()

    ACTION_ICON = {
        "CREATE":          "📦",
        "COMPLETE":        "✅",
        "DAMAGE":          "⚠️",
        "START":           "▶️",
        "UPDATE_DELIVERY": "🚚",
        "ADD_DELIVERY":    "📍",
        "ROUTE_PLAN":      "🗺️",
        "DELETE":          "🗑️",
        "LOGIN":           "🔐",
        "ASSIGN":          "👷",
        "STOCK_UPDATE":    "📊",
    }
    ACTION_LBL = {
        "CREATE":          "Oluşturuldu",
        "COMPLETE":        "Tamamlandı",
        "DAMAGE":          "Hasar Bildirimi",
        "START":           "Başlatıldı",
        "UPDATE_DELIVERY": "Teslimat Güncellendi",
        "ADD_DELIVERY":    "Teslimat Eklendi",
        "ROUTE_PLAN":      "Rota Planlandı",
        "DELETE":          "Silindi",
        "LOGIN":           "Giriş Yapıldı",
        "ASSIGN":          "Personel Atandı",
        "STOCK_UPDATE":    "Stok Güncellendi",
    }
    ENTITY_LBL = {
        "order":       "Sipariş",
        "order_item":  "Sipariş Kalemi",
        "deliveries":  "Teslimat",
        "route_plans": "Rota Planı",
        "pieces":      "Parça",
        "users":       "Kullanıcı",
        "piece_stock": "Stok",
    }
    ACTION_CLS = {
        "CREATE":          "card-blue",
        "COMPLETE":        "card-green",
        "DAMAGE":          "card-orange",
        "START":           "card-gray",
        "DELETE":          "card-red",
        "ROUTE_PLAN":      "card-blue",
        "UPDATE_DELIVERY": "card-blue",
        "ADD_DELIVERY":    "card-blue",
        "ASSIGN":          "card-gray",
        "STOCK_UPDATE":    "card-yellow",
    }

    if not logs:
        st.markdown("""
        <div style='text-align:center;padding:2.5rem;background:#161622;border-radius:12px;border:2px dashed #1e1e30'>
          <div style='font-size:2rem'>📋</div>
          <p style='color:#9ca3af;margin-top:.5rem'>Koşula uyan kayıt bulunamadı.</p>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    # Stats bar
    actions_count = {}
    for l in logs: actions_count[l["action"]] = actions_count.get(l["action"],0)+1
    cols_stat = st.columns(min(len(actions_count),5))
    for i,(act,cnt) in enumerate(sorted(actions_count.items(),key=lambda x:-x[1])[:5]):
        lbl = ACTION_LBL.get(act, act)
        icon = ACTION_ICON.get(act,"🔹")
        cols_stat[i].metric(f"{icon} {lbl}", cnt)

    st.divider()

    for log in logs:
        icon   = ACTION_ICON.get(log["action"],"🔹")
        cls    = ACTION_CLS.get(log["action"],"card-gray")
        act_lbl = ACTION_LBL.get(log["action"], log["action"])
        ent_lbl = ENTITY_LBL.get(log["entity"], log["entity"])
        st.markdown(f"""
        <div class='card {cls}' style='padding:.8rem 1.2rem;margin-bottom:.35rem'>
          <div style='display:flex;align-items:center;gap:.5rem;flex-wrap:wrap;margin-bottom:.3rem'>
            <span style='font-size:.95rem;font-weight:700'>{icon} {act_lbl}</span>
            <span class='badge badge-purple' style='font-size:.72rem'>{ent_lbl}</span>
            <span style='color:#6b7280;font-size:.78rem'>#{log["entity_id"] or "—"}</span>
          </div>
          <div style='font-size:.9rem;color:#d4d4e8;margin-bottom:.3rem'>{log["detail"] or "—"}</div>
          <div style='font-size:.78rem;color:#6b7280'>
            👤 {log["username"]} &nbsp;·&nbsp; 🕐 {log["created_at"]}
          </div>
        </div>
        """, unsafe_allow_html=True)

    # Export
    st.divider()
    if st.button("📊 CSV Olarak İndir", use_container_width=True):
        df_log = pd.DataFrame(logs)
        csv = df_log.to_csv(index=False, encoding="utf-8-sig")
        st.download_button("⬇️ Denetim Kaydı CSV", csv, "denetim_kaydi.csv", "text/csv",
                           use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
#  SAYFA: ADMİN PANELİ
# ══════════════════════════════════════════════════════════════════════════════
elif page=="⚙️ Admin Paneli":
    if not can("admin"): st.error("⛔ Yetki yok"); st.stop()
    st.title("⚙️ Admin Paneli")

    t1,t2,t3,t4,t5,t6 = st.tabs([
        "📋 Yükleme Sırası","🪑 Parçalar","👥 Kullanıcılar",
        "🚛 Araçlar","⏱ SLA Yapılandırma","🎲 Örnek Veri"
    ])

    # ── Yükleme Sırası ──
    with t1:
        colls_a=q("SELECT id,name FROM collections ORDER BY name")
        sel_ca=st.selectbox("Koleksiyon",[c["name"] for c in colls_a],key="ac")
        cid_a=next(c["id"] for c in colls_a if c["name"]==sel_ca)
        pcs_a=q("SELECT id,name,width_cm,depth_cm,height_cm,loading_order FROM pieces WHERE collection_id=? ORDER BY loading_order",(cid_a,))
        with st.form(f"of_{sel_ca}"):
            no={}
            for p in pcs_a:
                r1,r2,r3=st.columns([5,2,3])
                r1.markdown(f"**{p['name']}**")
                no[p["id"]]=r2.number_input("s",1,99,p["loading_order"],key=f"o_{p['id']}",label_visibility="collapsed")
                r3.caption(f"{p['width_cm']}×{p['depth_cm']}×{p['height_cm']} cm")
            if st.form_submit_button("💾 Kaydet",use_container_width=True):
                with get_conn() as c2:
                    for pid,v in no.items(): c2.execute("UPDATE pieces SET loading_order=? WHERE id=?",(v,pid))
                    c2.commit()
                st.success("✅ Güncellendi!"); st.rerun()

    # ── Parçalar ──
    with t2:
        colls_b=q("SELECT id,name FROM collections ORDER BY name")
        with st.form("ap"):
            st.markdown("**Yeni Parça Ekle**")
            sc=st.selectbox("Koleksiyon",[c["name"] for c in colls_b])
            pn=st.text_input("Parça Adı")
            b1,b2,b3=st.columns(3)
            pw2=b1.number_input("G (cm)",min_value=0.0); pd2=b2.number_input("D (cm)",min_value=0.0); ph2=b3.number_input("Y (cm)",min_value=0.0)
            po=st.number_input("Sıra",1,99,99)
            if st.form_submit_button("➕ Ekle",use_container_width=True):
                if not pn: st.error("Ad zorunlu")
                else:
                    pcid=next(c["id"] for c in colls_b if c["name"]==sc)
                    pid=ex_id("INSERT INTO pieces (collection_id,name,width_cm,depth_cm,height_cm,loading_order) VALUES(?,?,?,?,?,?)",(pcid,pn,pw2,pd2,ph2,po))
                    ex("INSERT OR IGNORE INTO piece_stock (piece_id) VALUES(?)",(pid,))
                    st.success(f"✅ '{pn}' eklendi!"); st.rerun()
        st.divider()
        dc=st.selectbox("Koleksiyon (sil)",  [c["name"] for c in colls_b],key="dc")
        dp=q("SELECT p.id,p.name FROM pieces p JOIN collections c ON p.collection_id=c.id WHERE c.name=? ORDER BY p.loading_order",(dc,))
        if dp:
            dm={p["name"]:p["id"] for p in dp}
            ds=st.selectbox("Parça",list(dm.keys()))
            if st.button("🗑️ Sil",type="primary"):
                ex("DELETE FROM pieces WHERE id=?",(dm[ds],)); st.success("✅ Silindi."); st.rerun()

    # ── Kullanıcılar ──
    with t3:
        users_l=q("SELECT id,username,role,full_name,is_active FROM users ORDER BY id")
        RL={"admin":"Admin","yonetici":"Yönetici","personel":"Personel"}
        for u2 in users_l:
            ab="badge-green" if u2["is_active"] else "badge-red"
            al="Aktif" if u2["is_active"] else "Pasif"
            rb={"admin":"badge-purple","yonetici":"badge-orange"}.get(u2["role"],"badge-gray")
            st.markdown(f"<div class='card card-gray' style='padding:.6rem 1.2rem;margin-bottom:.25rem'><b>{u2['full_name'] or u2['username']}</b> <span style='color:#9ca3af'>@{u2['username']}</span> &nbsp;<span class='badge {rb}'>{RL.get(u2['role'],u2['role'])}</span> &nbsp;<span class='badge {ab}'>{al}</span></div>",unsafe_allow_html=True)
        st.divider()
        with st.form("au"):
            st.markdown("**Yeni Kullanıcı**")
            nu1,nu2=st.columns(2); na=nu1.text_input("Kullanıcı Adı"); nfn=nu2.text_input("Ad Soyad")
            nu3,nu4=st.columns(2); np2=nu3.text_input("Şifre",type="password"); nr=nu4.selectbox("Rol",["personel","yonetici","admin"])
            if st.form_submit_button("➕ Ekle",use_container_width=True):
                if not na or not np2: st.error("Zorunlu alanlar eksik.")
                else:
                    try: ex("INSERT INTO users (username,password_hash,role,full_name) VALUES(?,?,?,?)",(na,sha256(np2),nr,nfn)); st.success("✅ Eklendi."); st.rerun()
                    except: st.error("Bu kullanıcı adı zaten var.")
        st.divider()
        tu=[u2 for u2 in users_l if u2["username"]!=user["username"]]
        if tu:
            tm={f"{u2['full_name'] or u2['username']} (@{u2['username']})":u2 for u2 in tu}
            ts=st.selectbox("Kullanıcı",list(tm.keys()),key="ts")
            tu2=tm[ts]; tl="Pasife Al" if tu2["is_active"] else "Aktife Al"
            if st.button(f"🔄 {tl}"):
                ex("UPDATE users SET is_active=? WHERE id=?",(0 if tu2["is_active"] else 1,tu2["id"])); st.rerun()

    # ── Araçlar ──
    with t4:
        vs2=q("SELECT id,name,capacity_m3 FROM vehicle_types ORDER BY capacity_m3 DESC")
        for v in vs2:
            st.markdown(f"<div class='card card-gray' style='padding:.6rem 1.2rem;margin-bottom:.25rem'><b>{v['name']}</b><span style='float:right;color:#a89bff;font-weight:700'>{v['capacity_m3']} m³</span></div>",unsafe_allow_html=True)
        st.divider()
        with st.form("av"):
            va,vb=st.columns(2); vn=va.text_input("Araç Adı"); vc=vb.number_input("Kapasite (m³)",min_value=.1,value=30.0)
            if st.form_submit_button("➕ Ekle",use_container_width=True):
                if not vn: st.error("Ad zorunlu")
                else:
                    try: ex("INSERT INTO vehicle_types (name,capacity_m3) VALUES(?,?)",(vn,vc)); st.rerun()
                    except: st.error("Bu araç tipi zaten var.")
        if vs2:
            vdm={v["name"]:v["id"] for v in vs2}; vds=st.selectbox("Sil",list(vdm.keys()),key="vds")
            if st.button("🗑️ Sil",type="primary",key="vdb"):
                ex("DELETE FROM vehicle_types WHERE id=?",(vdm[vds],)); st.rerun()

    # ── SLA Yapılandırma ──
    with t5:
        st.markdown("**Koleksiyon Bazında SLA Hedefleri (dakika)**")
        sla_rows=q("""SELECT c.name,sc.sla_minutes,c.id cid FROM sla_config sc
                      JOIN collections c ON sc.collection_id=c.id ORDER BY c.name""")
        with st.form("sla_form"):
            new_slas={}
            for sr in sla_rows:
                c1,c2=st.columns([3,2])
                c1.markdown(f"**{sr['name']}**")
                new_slas[sr["cid"]]=c2.number_input("dk",15,480,sr["sla_minutes"],key=f"sla_{sr['cid']}",label_visibility="collapsed")
            if st.form_submit_button("💾 SLA Kaydet",use_container_width=True):
                for cid2,v in new_slas.items():
                    ex("INSERT OR REPLACE INTO sla_config (collection_id,sla_minutes) VALUES(?,?)",(cid2,v))
                st.success("✅ SLA hedefleri güncellendi!"); st.rerun()

    # ── Örnek Veri ──
    with t6:
        st.markdown("### 🎲 Örnek Sipariş Üretici")
        with st.expander("ℹ️ Ürün Grubu Mantığı"):
            st.markdown("""
| Grup | İçerik | Dahil Olasılığı |
|---|---|---|
| 🛋️ Koltuk | Tüm koltuklar, berjer, tekli | %75 |
| 🍽️ Yemek Odası | Yemek masası + sandalye/puf | %60 |
| 🗄️ Depolama | Konsol + şifonyer + TV ünitesi | %55 |
| ☕ Sehpa | Orta sehpa + zigon sehpa + yan sehpa | %65 |
| 🪞 Aksesuar | Ayna vb. | %30 |
Bir grup seçilince gruptaki **tüm parçalar** birlikte gelir.
            """)
        n_ord=st.slider("Sipariş sayısı",5,50,20)
        c_pct=st.slider("Tamamlanmış oran (%)",20,90,65)
        st.caption(f"~{int(n_ord*c_pct/100)} tamamlanmış, ~{n_ord-int(n_ord*c_pct/100)} bekleyen")
        if st.button("🎲 Üret",use_container_width=True,type="primary"):
            from generate_sample_data import generate_orders,GROUP_LABELS,generate_all_extras
            with st.spinner("Oluşturuluyor..."):
                res=generate_orders(n=n_ord,completed_ratio=c_pct/100)
                extras=generate_all_extras(notif_n=12,audit_n=15,delivery_n=10)
            st.success(f"✅ {res['created']} sipariş oluşturuldu, {res['skipped']} atlandı.")
            st.caption(f"📬 {extras['notifications']} bildirim · 📜 {extras['audit_logs']} denetim kaydı · 📍 {extras['deliveries']} teslimat eklendi")
            if res["group_stats"]:
                mx=max(res["group_stats"].values()) or 1
                for g,cnt in sorted(res["group_stats"].items(),key=lambda x:-x[1]):
                    st.progress(cnt/mx,text=f"{GROUP_LABELS.get(g,g)}: {cnt}")
            st.rerun()
        st.divider()
        sm=q1("SELECT SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) p, SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) c FROM orders")
        if sm:
            c1,c2,c3=st.columns(3)
            c1.metric("Bekleyen",(sm["p"] or 0)); c2.metric("Tamamlanan",(sm["c"] or 0)); c3.metric("Toplam",(sm["p"] or 0)+(sm["c"] or 0))
        st.divider()
        with st.expander("⚠️ Tehlikeli Bölge"):
            st.warning("Tüm sipariş ve kalemleri siler. Geri alınamaz!")
            cf_check = st.checkbox("Evet, tüm siparişleri silmek istiyorum", key="cf_del_check")
            if st.button("🗑️ Tüm Siparişleri Sil", type="primary", key="dal", disabled=not cf_check):
                ex("DELETE FROM order_items"); ex("DELETE FROM orders")
                ex("DELETE FROM customer_tokens"); ex("DELETE FROM notifications")
                st.success("✅ Silindi."); st.rerun()

    st.divider()
    # ── E-posta Ayarları ──
    with st.expander("📧 E-posta (SMTP) Ayarları"):
        cfg_keys=["smtp_host","smtp_port","smtp_user","smtp_pass","company_name","app_url"]
        cfg_vals={r["key"]:r["value"] for r in q("SELECT key,value FROM app_settings")}
        with st.form("smtp_form"):
            sh=st.text_input("SMTP Host",    value=cfg_vals.get("smtp_host","smtp.gmail.com"))
            sp=st.text_input("SMTP Port",    value=cfg_vals.get("smtp_port","587"))
            su=st.text_input("SMTP User",    value=cfg_vals.get("smtp_user",""))
            spw=st.text_input("SMTP Şifre",  value=cfg_vals.get("smtp_pass",""),type="password")
            cn=st.text_input("Şirket Adı",   value=cfg_vals.get("company_name","Luxe Life Mobilya"))
            au=st.text_input("Uygulama URL", value=cfg_vals.get("app_url","http://localhost:8501"))
            if st.form_submit_button("💾 SMTP Ayarlarını Kaydet",use_container_width=True):
                for k,v in [("smtp_host",sh),("smtp_port",sp),("smtp_user",su),
                            ("smtp_pass",spw),("company_name",cn),("app_url",au)]:
                    ex("INSERT OR REPLACE INTO app_settings (key,value) VALUES(?,?)",(k,v))
                st.success("✅ Kaydedildi.")
